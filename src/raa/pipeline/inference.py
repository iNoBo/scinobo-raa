"""

SCINOBO RESEARCH ARTIFACT ANALYSIS TOOL

"""

""" 

*** RAE SYSTEM *** 

"""

""" IMPORTS """

import os
import gzip
import json
import fnmatch
import requests
import argparse
import importlib
from bs4 import BeautifulSoup
from nltk.tokenize import sent_tokenize
from collections import Counter, OrderedDict

from raa.pipeline.model import infer_text, classify_text

BASE_PATH = importlib.resources.files(__package__.split(".")[0])
DATA_PATH = os.path.join(BASE_PATH, 'data')

"""

COMPONENT 1 : Research Artifacts Gazetteers (RAG) 

This will be a dictionary that contains strings of research artifact names as keys and their types as a list of values.

e.g. "ARCH" : ["Dataset", "Method"]

"""

# We are going to get the methods and datasets from the paperswithcode.com dumps
with gzip.open(os.path.join(DATA_PATH, 'material/methods.json.gz'), 'r') as fin:
    pwc_methods = json.loads(fin.read().decode('utf-8'))

with gzip.open(os.path.join(DATA_PATH, 'material/datasets.json.gz'), 'r') as fin:
    pwc_datasets = json.loads(fin.read().decode('utf-8'))

# Flatten out the datasets and methods
pwc_flat_datasets = sorted(set([i.strip() for s in [[x['name']] + x['variants'] for x in pwc_datasets] for i in s if i]))
pwc_flat_methods = sorted(set([i.strip() for s in [[x['name'], x['full_name']] for x in pwc_methods] for i in s if i]))

# Create a dictionary of all the datasets and methods
research_artifact_gazetteers = dict()
for method in pwc_flat_methods:
    research_artifact_gazetteers[method] = ['Method']
for dataset in pwc_flat_datasets:
    if dataset not in research_artifact_gazetteers:
        research_artifact_gazetteers[dataset] = ['Dataset']
    else:
        research_artifact_gazetteers[dataset].append('Dataset')

# Remove some noise
del research_artifact_gazetteers['0']
del research_artifact_gazetteers['.']
del research_artifact_gazetteers['2048']  # TODO-CHECK: this is the 2048 game, should we remove it?

""" 

COMPONENT 2 : PDF Parser 

We are going to use a local GROBID server to parse the pdfs. 

The GROBID server is a REST API that takes in a pdf and returns a TEI XML file that contains the sections of the pdf.

The TEI XML contains the following information:
- paper metadata (authors, title, abstract, affiliations etc)
- the sections of the paper
- the paragraphs of the paper
- the sentences of the paper
- bibliographical references + citations
- other citations (to formulas, figures, etc) : TODO-CHECK
- etc

SEE ALSO:
- https://gitlab.com/internetarchive/grobid_tei_xml
- https://github.com/allenai/s2orc-doc2json
- https://github.com/delb-xml/delb-py
- https://komax.github.io/blog/text/python/xml/parsing_tei_xml_python/

"""

grobid_url = os.getenv('GROBID_URL')
if not grobid_url:
    grobid_url = 'https://kermitt2-grobid.hf.space/api/processFulltextDocument'
else:
    grobid_url = f'{grobid_url}/api/processFulltextDocument'


def parse_pdf(pdf_path):
    """
    > node names in the TEI XML file:
        ['TEI', 'abstract', 'affiliation', 'analytic', 'appInfo', 'application', 'author', 'availability', 'back', 'biblScope', 'biblStruct', 
        'body', 'cell', 'date', 'desc', 'div', 'email', 'encodingDesc', 'facsimile', 'figDesc', 'figure', 'fileDesc', 'forename', 'formula', 
        'graphic', 'head', 'idno', 'imprint', 'label', 'licence', 'listBibl', 'monogr', 'note', 'orgName', 'p', 'persName', 'profileDesc', 
        'pubPlace', 'publicationStmt', 'publisher', 'ref', 'respStmt', 'row', 's', 'sourceDesc', 'surface', 'surname', 'table', 'teiHeader', 
        'text', 'title', 'titleStmt']
    
    """
    with open(pdf_path, 'rb') as fin:
        pdf_data = fin.read()
    
    payload = {
        'input': pdf_data,
        'consolidateHeader': 1,
        'consolidateCitations': 1,
        'includeRawCitations': 1,
        'includeRawAffiliations': 1,
        'teiCoordinates': 1,
        'segmentSentences': 1
    }

    res = requests.post(grobid_url, files=payload)
    
    pdf_metadata = parse_tei_xml(res.text, is_file=False)

    return pdf_metadata


def parse_tei_xml(tei_xml, is_file=True):
    """
    > node names in the TEI XML file:
        ['TEI', 'abstract', 'affiliation', 'analytic', 'appInfo', 'application', 'author', 'availability', 'back', 'biblScope', 'biblStruct', 
        'body', 'cell', 'date', 'desc', 'div', 'email', 'encodingDesc', 'facsimile', 'figDesc', 'figure', 'fileDesc', 'forename', 'formula', 
        'graphic', 'head', 'idno', 'imprint', 'label', 'licence', 'listBibl', 'monogr', 'note', 'orgName', 'p', 'persName', 'profileDesc', 
        'pubPlace', 'publicationStmt', 'publisher', 'ref', 'respStmt', 'row', 's', 'sourceDesc', 'surface', 'surname', 'table', 'teiHeader', 
        'text', 'title', 'titleStmt']
    
    """

    if is_file:
        # Load the TEI XML file
        with open(tei_xml, 'r', encoding='utf-8') as fin:
            tei_xml_data = fin.read()
    else:
        # The function has TEI XML data directly
        tei_xml_data = tei_xml

    soup = BeautifulSoup(tei_xml_data, 'xml')

    pdf_metadata = dict()

    if soup.find('sourceDesc'):
        identifiers = [[x.attrs['type'], x.text] for x in soup.find('sourceDesc').find_all('idno')]
    else:
        identifiers = None
    pdf_metadata['identifiers'] = identifiers if identifiers else None

    title = soup.find('title')
    pdf_metadata['title'] = title.text if title else None

    abstract = soup.find('abstract')
    abstract = [[s.text for s in p.find_all('s')] for p in abstract.find_all('p')]  # find all paragraphs and then all sentences
    pdf_metadata['abstract'] = abstract if abstract else None

    sections = soup.find_all('div')
    # For each section, find the head (if there is one), all paragraphs and then all sentences
    sections = [[s.head.text if s.head else None, [[[s, s.text] for s in p.find_all('s')] for p in s.find_all('p')]] for s in sections]

    # When a section has no head, then search if it is a figure, table or graphic using the "find_parent" method
    for i, section in enumerate(sections):
        if (section[0] is None) and (len(section[1])!=0) and (section[1][0] != []):
            sec_first_s = section[1][0][0][0]
            if sec_first_s.find_parent('figure'):
                par = sec_first_s.find_parent('figure')
                sections[i][0] = par.find('head').text if par.find('head') else 'figure'
            elif sec_first_s.find_parent('table'):
                par = sec_first_s.find_parent('table')
                sections[i][0] = par.find('head').text if par.find('head') else 'table'
            elif sec_first_s.find_parent('graphic'):
                par = sec_first_s.find_parent('graphic')
                sections[i][0] = par.find('head').text if par.find('head') else 'graphic'

    # Remove the extra from the sentences
    for i, section in enumerate(sections):
        for j, paragraph in enumerate(section[1]):
            for k, sentence in enumerate(paragraph):
                sections[i][1][j][k] = sentence[1]
    
    pdf_metadata['sections'] = sections

    # Find the references in the bibliography
    bibl_references = [{
        'id': r.attrs['xml:id'] if 'xml:id' in r.attrs else None,
        'doi': r.find('idno', {'type': 'DOI'}).text.lower() if r.find('idno', {'type': 'DOI'}) else None,
        'title': r.find('analytic').find('title').text if r.find('analytic') and r.find('analytic').find('title') else None,
        'authors': [{
            'surname': a.find('surname').text if a.find('surname') else None,
            'forename': a.find('forename').text if a.find('forename') else None,
            # 'email': a.find('email').text if a.find('email') else None,
            # 'affiliation': {
            #     'id': a.find('affiliation').attrs['key'] if 'key' in a.find('affiliation').attrs else None,
            #     'name': a.find('orgName').text if a.find('orgName') else None,
            # } if a.find('affiliation') else None
            } for a in r.find('analytic').find_all('author')] if r.find('analytic') else None,
        } for r in soup.find_all('biblStruct')]

    # Find the figures in the text
    figures = [{
        'id': x.attrs['xml:id'] if 'xml:id' in x.attrs else None, 
        'head': x.find('head').text if x.find('head') else None,
        'desc': x.find('figDesc').text if x.find('figDesc') else None
        } for x in soup.find_all('figure')]

    # Find the formulas in the text
    formulas = [{
        'id': x.attrs['xml:id'] if 'xml:id' in x.attrs else None,
        'desc': x.text,
        } for x in soup.find_all('formula')]


    # Find the citances in the section, paragraph and sentence
    citances = [[[[sec_i, par_i, s_i, s.find_all('ref')] for s_i, s in enumerate(par.find_all('s')) if s.find('ref')] for par_i, par in enumerate(sec.find_all('p')) if par.find('ref')] for sec_i, sec in enumerate(soup.find_all('div')) if sec.find('ref')]
    citances = [i2 for s2 in [i for s in citances for i in s] for i2 in s2]
    citances = [{
            'sec_idx': x[0], 
            'par_idx': x[1], 
            's_idx': x[2], 
            'refs':
                [{
                'target': y.attrs['target'] if 'target' in y.attrs else None,
                'type': y.attrs['type'] if 'type' in y.attrs else None,
                'text': y.text,} for y in x[3]
                ],
            'sentence': sections[x[0]][1][x[1]][x[2]]} for x in citances]

    pdf_metadata['bibl_references'] = bibl_references
    pdf_metadata['figures'] = figures
    pdf_metadata['formulas'] = formulas
    pdf_metadata['citances'] = citances

    return pdf_metadata

""" 

COMPONENT 3 : Candidate Detection System (CDS) 

We are going to use dataset and software keywords to detect candidate research artifacts in the pdfs.

Another approach would be to use a Sequence Tagging model to predict the candidates.

TODO-CHECK: Maybe we could also use the gazeetters from the previous component to detect candidates.

"""

import re
from tqdm import tqdm

with open(os.path.join(DATA_PATH, 'material/data_phrases.txt'), 'r') as fin:
    data_phrases = [x.strip() for x in fin.readlines()]

with open(os.path.join(DATA_PATH, 'material/software_phrases.txt'), 'r') as fin:
    software_phrases = [x.strip() for x in fin.readlines()]

# def my_contains_func(sentence, phrase):
#     if phrase in sentence:
#         return True
#     return False


def my_contains_func(sentence, phrase, return_offsets=True, ignore_case=True):
    if ignore_case:
        sentence = sentence.lower()
        phrase = phrase.lower()
    if phrase in sentence:
        if return_offsets:
            ph_match = [x.span() for x in re.finditer(r'\b' + re.escape(phrase) + r'\b', sentence)]
            return ph_match
        else:
            return True
    if return_offsets:
        return []
    else:
        return False


def detect_candidates(pdf_metadata, gazetteers, external_gazetteers, data_phrases, software_phrases, contains_func=my_contains_func, gaz_contains_func=my_contains_func, verbose=True):
    """
    INPUTS
    - pdf_metadata: the metadata of the pdf (title, abstract, sections)
    - gazetteers: the gazetteers (datasets, methods) currently from PWC
    - external_gazetteers: gazetteers (datasets, methods) from external sources
    - data_phrases: the phrases that indicate a dataset
    - software_phrases: the phrases that indicate a software
    - contains_func: the function that checks if a phrase is in a sentence
    - gaz_contains_func: the function that checks if a gazetteer is in a sentence

    RETURNS
    - candidates: a dictionary with the [section, paragraph, sentence, candidate_offsets] of each candidates from each type (dataset, software, gaz_dataset, gaz_method)

    """
    candidates = {
        'dataset': dict(),
        'software': dict(),
        'gaz_dataset': dict(),
        'gaz_method': dict()
    }
    t_pdf_metadata_sections = tqdm(pdf_metadata['sections']) if verbose else pdf_metadata['sections']
    for i, section in enumerate(t_pdf_metadata_sections):
        for j, paragraph in enumerate(section[1]):
            if verbose:
                t_pdf_metadata_sections.set_description(f'Processing paragraph {j+1}/{len(section[1])}')
            for k, sentence in enumerate(paragraph):
                # Detect gazeetters
                for gazetteer in gazetteers:
                    gaz_matches = gaz_contains_func(sentence, gazetteer, ignore_case=True)  # B Version
                    if gaz_matches:
                        for gaz_match in gaz_matches:
                            gaz_types = gazetteers[gazetteer]
                            for g_type in gaz_types:
                                if g_type == 'Dataset':
                                    g_type = 'gaz_dataset'
                                elif g_type == 'Method':
                                    g_type = 'gaz_method'
                                else:
                                    raise Exception(f'Unknown gazetteer type: {g_type}')
                                if gazetteer not in candidates[g_type]:
                                    candidates[g_type][gazetteer] = [[i, j, k, gaz_match]]
                                else:
                                    candidates[g_type][gazetteer].append([i, j, k, gaz_match])
                for gazetteer in external_gazetteers:
                    gaz_matches = gaz_contains_func(sentence, gazetteer, ignore_case=True)  # B Version
                    if gaz_matches:
                        for gaz_match in gaz_matches:
                            gaz_types = external_gazetteers[gazetteer]
                            for g_type in gaz_types:
                                if g_type == 'Dataset':
                                    g_type = 'gaz_dataset'
                                elif g_type == 'Method':
                                    g_type = 'gaz_method'
                                else:
                                    raise Exception(f'Unknown gazetteer type: {g_type}')
                                if gazetteer not in candidates[g_type]:
                                    candidates[g_type][gazetteer] = [[i, j, k, gaz_match]]
                                else:
                                    candidates[g_type][gazetteer].append([i, j, k, gaz_match])
                # Detect data phrases
                for data_phrase in data_phrases:
                    ph_matches = contains_func(sentence, data_phrase, ignore_case=True)
                    if ph_matches:
                        for ph_match in ph_matches:
                            if data_phrase not in candidates['dataset']:
                                candidates['dataset'][data_phrase] = [[i, j, k, ph_match]]
                            else:
                                candidates['dataset'][data_phrase].append([i, j, k, ph_match])
                # Detect software phrases
                for software_phrase in software_phrases:
                    ph_matches = contains_func(sentence, software_phrase)
                    if ph_matches:
                        for ph_match in ph_matches:
                            if software_phrase not in candidates['software']:
                                candidates['software'][software_phrase] = [[i, j, k, ph_match]]
                            else:
                                candidates['software'][software_phrase].append([i, j, k, ph_match])
    return candidates


def detect_candidates_list(text_list, gazetteers, external_gazetteers, data_phrases, software_phrases, contains_func=my_contains_func, gaz_contains_func=my_contains_func, verbose=True):
    """
    INPUTS
    - text_list: the list of lists of text to find the datasets (paragraphs, sentences)
    - gazetteers: the gazetteers (datasets, methods) currently from PWC
    - external_gazetteers: gazetteers (datasets, methods) from external sources
    - data_phrases: the phrases that indicate a dataset
    - software_phrases: the phrases that indicate a software
    - contains_func: the function that checks if a phrase is in a sentence
    - gaz_contains_func: the function that checks if a gazetteer is in a sentence

    RETURNS
    - candidates: a dictionary with the [section, paragraph, sentence, candidate_offsets] of each candidates from each type (dataset, software, gaz_dataset, gaz_method)

    """
    candidates = {
        'dataset': dict(),
        'software': dict(),
        'gaz_dataset': dict(),
        'gaz_method': dict()
    }
    for p, paragraph in enumerate(text_list):
        for k, sentence in enumerate(paragraph):
            # Detect gazeetters
            for gazetteer in gazetteers:
                gaz_matches = gaz_contains_func(sentence, gazetteer, ignore_case=True)  # B version
                if gaz_matches:
                    for gaz_match in gaz_matches:
                        gaz_types = gazetteers[gazetteer]
                        for g_type in gaz_types:
                            if g_type == 'Dataset':
                                g_type = 'gaz_dataset'
                            elif g_type == 'Method':
                                g_type = 'gaz_method'
                            else:
                                raise Exception(f'Unknown gazetteer type: {g_type}')
                            if gazetteer not in candidates[g_type]:
                                candidates[g_type][gazetteer] = [[p, k, gaz_match]]
                            else:
                                candidates[g_type][gazetteer].append([p, k, gaz_match])
            for gazetteer in external_gazetteers:
                gaz_matches = gaz_contains_func(sentence, gazetteer, ignore_case=True)  # B version
                if gaz_matches:
                    for gaz_match in gaz_matches:
                        gaz_types = external_gazetteers[gazetteer]
                        for g_type in gaz_types:
                            if g_type == 'Dataset':
                                g_type = 'gaz_dataset'
                            elif g_type == 'Method':
                                g_type = 'gaz_method'
                            else:
                                raise Exception(f'Unknown gazetteer type: {g_type}')
                            if gazetteer not in candidates[g_type]:
                                candidates[g_type][gazetteer] = [[p, k, gaz_match]]
                            else:
                                candidates[g_type][gazetteer].append([p, k, gaz_match])
            # Detect data phrases
            for data_phrase in data_phrases:
                ph_matches = contains_func(sentence, data_phrase, ignore_case=True)
                if ph_matches:
                    for ph_match in ph_matches:
                        if data_phrase not in candidates['dataset']:
                            candidates['dataset'][data_phrase] = [[p, k, ph_match]]
                        else:
                            candidates['dataset'][data_phrase].append([p, k, ph_match])
            # Detect software phrases
            for software_phrase in software_phrases:
                ph_matches = contains_func(sentence, software_phrase)
                if ph_matches:
                    for ph_match in ph_matches:
                        if software_phrase not in candidates['software']:
                            candidates['software'][software_phrase] = [[p, k, ph_match]]
                        else:
                            candidates['software'][software_phrase].append([p, k, ph_match])
    return candidates

""" 

COMPONENT 4 : Research Artifact eXtractor (RAX) 

We are going to PRIMER OUGHT and ICE to extract the research artifacts from the pdfs using LLMs (GPT3, etc) with the following recipe:
- we will create a table that contains all candidate sentences and their paragraphs and sections
- for each candidate sentence, we will flag which contain citation numbers
- we can also use correference resolution to find which sentences are referring to the same research artifact (using hugginface) : TODO-LATER
 - create correference chains
- for each candidate sentence we are going to ask the GPT3 model the following questions:
 - what is the name of the research artifact?
 - what is the license of the research artifact?
 - what is the version to the research artifact?
 - what is the url to the research artifact?
 - what is the description of the research artifact?
 - what is the citation of the research artifact?
 - is the research artifact created by the authors?
 - is the research artifact used by the authros?
 --> we add this info to the table of candidates
 --> if some question require more text, then we give the paragraph and the section to the GPT3 model
- for each candidate we determine whether it is a new research artifact or not
 - we check using the GPT3 model if the research artifact is already in the cache of research artifacts
 - if it is not, then we add it to the cache of research artifacts for the paper

"""
import requests

# Initialize thresholds
thresholds = {
    'paragraph_answer': 0.0,
    'artifact_answer': 0.0,
    # 'name': 0.0,
    # 'license': 0.0,
    # 'version': 0.0,
    'ownership_answer': 0.0,
    'reuse_answer': 0.0,
    # 'cache_answer': 0.9,
}

# FOR COREFERENCE RESOLUTION (via API)
import requests

scico_api_url = os.getenv('SCICO_API_URL')
if str(scico_api_url) == 'None':
    scico_api_url = None
    
if not scico_api_url:
    print('SCICO_API_URL not set, skipping coreference resolution in deduplication')
    scico_api_url = 'SKIP'


def check_coreference_resolution(m1, m2):
    if scico_api_url == 'SKIP':
        return {'not related': 1.0, 'corefer': 0.0, 'parent': 0.0, 'child': 0.0}
    res = requests.post('{}/scico_infer'.format(scico_api_url), json={'m1': m1, 'm2': m2}).json()
    return res

def check_coreference_resolution_batch(batch):
    if scico_api_url == 'SKIP':
        return [{'not related': 1.0, 'corefer': 0.0, 'parent': 0.0, 'child': 0.0} for _ in batch]
    res = requests.post('{}/scico_infer_batch'.format(scico_api_url), json={'batch': batch}).json()
    return res

# IDEA CONCEPT 3
# 1. first cluster based on the name
# 2. then use the scico model to cluster all named clusters by themselves
# 3. then cluster N/A mentions to named clusters (only in the same paragraph)
# 4. then cluster N/A mentions by themselves (only in the same paragraph)
# 5. then cluster unnamed clusters to the named clusters (in any context)
# 6. then merge the clusters that have named clusters with the same citation marks
def deduplicate_mentions_dedup(mentions):
    def create_text_annotations(text, pattern):
        matches = re.finditer(re.escape(pattern), text)
        annotated_texts = []

        for match in matches:
            start, end = match.span()  # group(0) is the whole match
            annotated_text = text[:start] + '<m>' + text[start:end] + '</m>' + text[end:]
            annotated_texts.append(annotated_text)

        return annotated_texts
    

    def check_similarity(mention, cluster, batched=True):
        if batched:
            batches = []
            for m in cluster:
                batches.append((mention[2], m[2]))
            # Max batch size
            b_size = 8
            batches = [batches[i:i + b_size] for i in range(0, len(batches), b_size)]
            score_list = []
            for batch in batches:
                scores = check_coreference_resolution_batch(batch)
                for s in scores:
                    score_list.append(s['corefer'] + s['parent'] + s['child'])
            average_score = np.mean(score_list)
        else:
            score_list = []
            for m in cluster:
                res = check_coreference_resolution(mention[2], m[2])
                score = res['corefer'] + res['parent'] + res['child']
                score_list.append(score)
            average_score = np.mean(score_list)
        return average_score


    def check_similarity_clusters(cluster1, cluster2, batched=True):
        if batched:
            batches = []
            for m1 in cluster1:
                for m2 in cluster2:
                    batches.append((m1[2], m2[2]))
            # Max batch size
            b_size = 8
            batches = [batches[i:i + b_size] for i in range(0, len(batches), b_size)]
            score_list = []
            for batch in batches:
                scores = check_coreference_resolution_batch(batch)
                for s in scores:
                    score_list.append(s['corefer'] + s['parent'] + s['child'])
            average_score = np.mean(score_list)
        else:
            score_list = []
            for m1 in cluster1:
                for m2 in cluster2:
                    res = check_coreference_resolution(m1[2], m2[2])
                    score = res['corefer'] + res['parent'] + res['child']
                    score_list.append(score)
            average_score = np.mean(score_list)
        return average_score


    import numpy as np
    from tqdm import trange
    from sklearn.cluster import AgglomerativeClustering

    # Step 0: Cluster the mentions based on their paragraph
    paragraph_clusters = {}
    for m in mentions:
        m_key = (m[1]['indices'][0], m[1]['indices'][1])
        if m_key not in paragraph_clusters:
            paragraph_clusters[m_key] = [m]
        else:
            paragraph_clusters[m_key].append(m)

    # Step 1: Cluster mention based on the name (also using the name_groups)
    name_clusters = {}
    for mention in tqdm(mentions):
        name = mention[1]['results']['name_answer']
        if name == 'N/A':
            continue
        if name not in name_clusters:
            name_clusters[name] = [mention]
        else:
            name_clusters[name].append(mention)

    name_clusters_paragraphs = {k: set([(m[1]['indices'][0], m[1]['indices'][1]) for m in name_clusters[k]]) for k in name_clusters}

    # Step 1b: Cluster mentions of different names using scico model
    if len(name_clusters) > 1:
        similarity_matrix = np.zeros((len(name_clusters), len(name_clusters)))
        name_clusters_list = list(name_clusters)
        t_name_clusters_list = trange(len(name_clusters_list))
        for i in t_name_clusters_list:
            similarity_matrix[i][i] = 1.0
            for j in range(i+1, len(name_clusters_list)):
                # Show the progress for j
                t_name_clusters_list.set_description(f'Processing {j}/{len(name_clusters_list)}')

                # # Using the triggers of the mentions
                # i_mentions = [x[2] for x in name_clusters[name_clusters_list[i]]]
                # j_mentions = [x[2] for x in name_clusters[name_clusters_list[j]]]

                # Using the normalized name of the mentions
                i_mentions = list()
                for m in name_clusters[name_clusters_list[i]]:
                    # par_pre = m[1]['paragraph'].split(m[1]['snippet'])[0]
                    # par_post = m[1]['paragraph'].split(m[1]['snippet'])[1]
                    par_pre = ''
                    par_post = ''
                    i_mentions.extend([par_pre + x + par_post for x in create_text_annotations(m[1]['snippet'], re.escape(m[1]['results']['name_answer']))])
                j_mentions = list()
                for m in name_clusters[name_clusters_list[j]]:
                    # par_pre = m[1]['paragraph'].split(m[1]['snippet'])[0]
                    # par_post = m[1]['paragraph'].split(m[1]['snippet'])[1]
                    par_pre = ''
                    par_post = ''
                    j_mentions.extend([par_pre + x + par_post for x in create_text_annotations(m[1]['snippet'], re.escape(m[1]['results']['name_answer']))])

                # Find the average similarity between the two clusters
                total_similarity = 0
                my_total = 0
                
                # BATCHED
                batches = list()
                for i_m in i_mentions:
                    for j_m in j_mentions:
                        batches.append((i_m, j_m))
                # Max batch size
                b_size = 8
                batches = [batches[i:i + b_size] for i in range(0, len(batches), b_size)]
                for batch in batches:
                    scores = check_coreference_resolution_batch(batch)
                    for score in scores:
                        total_similarity += score['corefer'] + score['parent'] + score['child']
                        my_total += 1

                if my_total == 0:
                    similarity_matrix[i][j] = 0.0
                    similarity_matrix[j][i] = 0.0
                else:
                    similarity_matrix[i][j] = total_similarity / my_total
                    similarity_matrix[j][i] = similarity_matrix[i][j]
        
        distance_matrix = 1 - similarity_matrix
        # Cluster the name cluster groups
        grouped_name_clusters = {}
        try:
            clustering = AgglomerativeClustering(n_clusters=None, distance_threshold=0.5, metric='precomputed', linkage='average').fit(distance_matrix)
        except TypeError:
            clustering = AgglomerativeClustering(n_clusters=None, distance_threshold=0.5, affinity='precomputed', linkage='average').fit(distance_matrix)

        for i, cluster in enumerate(clustering.labels_):
            cluster_name = f'name_cluster_{cluster}'
            if cluster_name not in grouped_name_clusters:
                grouped_name_clusters[cluster_name] = [name_clusters_list[i]]
            else:
                grouped_name_clusters[cluster_name].append(name_clusters_list[i])
    else:
        grouped_name_clusters = {'name_cluster_0': list(name_clusters)}

    # Step 2: Use the scico model to cluster N/A mentions to named clusters (only in the same paragraph)
    n_a_clusters = []
    for mention in tqdm(mentions):
        mention_paragraph_index = (mention[1]['indices'][0], mention[1]['indices'][1])
        name = mention[1]['results']['name_answer']
        if name != 'N/A':
            continue
        similarity_scores = []
        name_clusters_filtered = {k: name_clusters[k] for k in name_clusters if mention_paragraph_index in name_clusters_paragraphs[k]}
        for named_cluster in name_clusters_filtered.values():
            similarity_scores.append(check_similarity(mention, named_cluster))
        if similarity_scores:
            closest_cluster = np.argmax(similarity_scores)
            if similarity_scores[closest_cluster] > 0.5:
                name_clusters[list(name_clusters_filtered.keys())[closest_cluster]].append(mention)
            else:
                n_a_clusters.append(mention)
        else:
            n_a_clusters.append(mention)
        
    # Step 3: If the N/A mentions are not that similar to the named clusters, then try to cluster them by themselves (only in the same paragraph)
    if n_a_clusters:
        if len(n_a_clusters) > 1:
            # Find the similarity matrix for the N/A clusters
            similarity_matrix = np.zeros((len(n_a_clusters), len(n_a_clusters)))
            for i in range(len(n_a_clusters)):
                i_paragraph_index = (n_a_clusters[i][1]['indices'][0], n_a_clusters[i][1]['indices'][1])
                similarity_matrix[i][i] = 1
                for j in range(i+1, len(n_a_clusters)):
                    j_paragraph_index = (n_a_clusters[j][1]['indices'][0], n_a_clusters[j][1]['indices'][1])
                    if i_paragraph_index != j_paragraph_index:
                        continue
                    res = check_coreference_resolution(n_a_clusters[i][2], n_a_clusters[j][2])
                    try:
                        similarity_matrix[i][j] = res['corefer'] + res['parent'] + res['child']
                    except:
                        raise Exception(f'Error in the scico model: {res}')
                    similarity_matrix[j][i] = similarity_matrix[i][j]
            # Calculate the distance matrix
            distance_matrix = 1 - similarity_matrix
            # Cluster them
            try:
                clustering = AgglomerativeClustering(n_clusters=None, distance_threshold=0.5, metric='precomputed', linkage='average').fit(distance_matrix)
            except TypeError:
                clustering = AgglomerativeClustering(n_clusters=None, distance_threshold=0.5, affinity='precomputed', linkage='average').fit(distance_matrix)

            for i, cluster in enumerate(clustering.labels_):
                cluster_fake_name = 'Unnamed_{}'.format(cluster)
                if cluster_fake_name not in name_clusters:
                    name_clusters[cluster_fake_name] = [n_a_clusters[i]]
                else:
                    name_clusters[cluster_fake_name].append(n_a_clusters[i])
        else:
            name_clusters['Unnamed_0'] = n_a_clusters

    # Step 4: Cluster unnamed clusters to the named clusters (in any context)
    unnamed_clusters = [k for k in name_clusters if k.startswith('Unnamed')]
    named_clusters = [k for k in name_clusters if not k.startswith('Unnamed')]
    for cluster1 in tqdm(unnamed_clusters):
            similarity_scores = []
            for cluster2 in named_clusters:
                similarity_scores.append(check_similarity_clusters(name_clusters[cluster1], name_clusters[cluster2]))
            if similarity_scores:
                best_cluster = np.argmax(similarity_scores)
                if similarity_scores[best_cluster] > 0.5:
                    name_clusters[list(name_clusters.keys())[best_cluster]].extend(name_clusters[cluster1])
                    del name_clusters[cluster1]

    # Step 5: Rename all unnamed clusters from 0
    to_rename = dict()
    un_c = 0
    for cluster in name_clusters:
        if cluster.startswith('Unnamed'):
            to_rename[cluster] = 'Unnamed_{}'.format(un_c)
            un_c += 1
    for k in to_rename:
        name_clusters[to_rename[k]] = name_clusters[k]
        del name_clusters[k]
    
    # Step 6: Create the final clusters
    grouped_name_clusters = {g_k: {k: name_clusters[k] for k in grouped_name_clusters[g_k]} for g_k in grouped_name_clusters}
    # Add the unnamed clusters to the final clusters
    for k in name_clusters:
        if k.startswith('Unnamed'):
            if 'Unnamed' not in grouped_name_clusters:
                grouped_name_clusters['Unnamed'] = {k: name_clusters[k]}
            else:
                grouped_name_clusters['Unnamed'][k] = name_clusters[k]
    
    # Step 7: Perform a post-processing of the deduplication function to try to merge the clusters that have the same citation marks

    # Find the closest citation mark for each mention in each cluster (using the span of the cluster name)
    all_artifact_names = [i for s in [list(grouped_name_clusters[k]) for k in grouped_name_clusters if k!='Unnamed'] for i in s]
    for g_cluster in grouped_name_clusters:
        for n_cluster in grouped_name_clusters[g_cluster]:
            for mention in grouped_name_clusters[g_cluster][n_cluster]:
                mention_citations = set([i for s in mention[1]['citations'] for i in s])
                if mention_citations:
                    mention_snippet = mention[1]['snippet']
                    # Find all spans of the cluster name in the snippet
                    name_spans = [[m.start(), m.end()] for m in re.finditer(re.escape(n_cluster), mention_snippet)]
                    if name_spans:
                        # Find all spans for the rest of the artifact names in the snippet
                        # NOTE: we want to avoid assigning the closest citation to the name span of another artifact
                        other_name_spans = list()
                        for o_name in all_artifact_names:
                            if o_name != n_cluster:
                                other_name_spans.extend([[m.start(), m.end()] for m in re.finditer(re.escape(o_name), mention_snippet)])
                        # Find the closest citation mark to the name span
                        closest_citation = None
                        closest_citation_span = None
                        closest_citation_distance = float('inf')
                        for citation in mention_citations:
                            citation_spans = [[m.start(), m.end()] for m in re.finditer(re.escape(citation), mention_snippet)]
                            if citation_spans:
                                # Find the span that is closest to the name span
                                for citation_span in citation_spans:
                                    for name_span in name_spans:
                                        if abs(citation_span[0] - name_span[0]) < closest_citation_distance:
                                            closest_citation_distance = abs(citation_span[0] - name_span[0])
                                            closest_citation_span = citation_span
                                            closest_citation = citation
                        # Keep the closest citation only if there are no other name spans closer to it
                        if other_name_spans and closest_citation is not None:
                            for o_name_span in other_name_spans:
                                if o_name_span != closest_citation_span:
                                    # If the other name span is in the citation span, then we keep the citation
                                    if o_name_span[0] >= closest_citation_span[0] and o_name_span[1] <= closest_citation_span[1]:
                                        continue
                                    # If the other name span is closer to the citation span, then we remove the citation
                                    if abs(closest_citation_span[0] - o_name_span[0]) < closest_citation_distance:
                                        closest_citation = None
                                        break
                        mention[1]['closest_citation'] = closest_citation
                    else:
                        mention[1]['closest_citation'] = None
                else:
                    mention[1]['closest_citation'] = None

    # Find the clusters that have the same closest citation mark

    # NEW WAY: each name cluster must have only one closest citation mark
    artifacts_closest_citations = [[i[0], i[1][0][0] if i[1] else ''] for s in [[[x, Counter([x2[1]['closest_citation'] for x2 in grouped_name_clusters[k][x] if x2[1]['closest_citation']]).most_common(1)] for x in grouped_name_clusters[k]] for k in grouped_name_clusters] for i in s]
    artifacts_closest_citations = {k: v for k, v in artifacts_closest_citations if v}

    # Merge the clusters that have the same closest citation mark using the artifacts closest citation mark
    clusters_to_merge = []
    for g_cluster in grouped_name_clusters:
        for n_cluster in grouped_name_clusters[g_cluster]:
            for g_cluster2 in grouped_name_clusters:
                if g_cluster != g_cluster2:
                    for n_cluster2 in grouped_name_clusters[g_cluster2]:
                        if n_cluster != n_cluster2:
                            try:
                                if artifacts_closest_citations[n_cluster] == artifacts_closest_citations[n_cluster2]:
                                    clusters_to_merge.append([g_cluster, n_cluster, g_cluster2, n_cluster2])
                            except KeyError:
                                continue

    # Keep only the "grouped name clusters" to merge
    grouped_clusters_to_merge = set([tuple(sorted([x[0], x[2]], key=lambda y: int(y.split('_')[-1]))) for x in clusters_to_merge])
    # Create sets of the clusters to merge
    def find_common_elements(tuple1, tuple2):
        return set(tuple1).intersection(set(tuple2))

    def merge_tuples(tuple1, tuple2):
        return tuple(set(tuple1).union(set(tuple2)))

    def merge_common(clustered_tuples):
        new_clustered_tuples = []
        merged_indexes = set()
        for i, t1 in enumerate(clustered_tuples):
            if i in merged_indexes:
                continue
            merged = False
            for j, t2 in enumerate(clustered_tuples[i+1:], i+1):
                if j in merged_indexes:
                    continue
                if find_common_elements(t1, t2):
                    new_clustered_tuples.append(merge_tuples(t1, t2))
                    merged_indexes.add(i)
                    merged_indexes.add(j)
                    merged = True
                    break
            if not merged:
                new_clustered_tuples.append(t1)
        return new_clustered_tuples

    def cluster_tuples(tuples_list):
        clustered_tuples = list(tuples_list)
        while True:
            new_clustered_tuples = merge_common(clustered_tuples)
            if len(new_clustered_tuples) == len(clustered_tuples):
                break
            clustered_tuples = new_clustered_tuples
        return clustered_tuples
    
    grouped_clusters_to_merge = cluster_tuples(grouped_clusters_to_merge)
    
    # Merge the clusters
    for cluster_group in grouped_clusters_to_merge:
        new_cluster_name = 'name_cluster_{}'.format('_'.join([x.split('_')[-1] for x in cluster_group]))
        grouped_name_clusters[new_cluster_name] = {}
        for cluster in cluster_group:
            grouped_name_clusters[new_cluster_name].update(grouped_name_clusters[cluster])
            del grouped_name_clusters[cluster]
    
    # Step 8 : Find abbreviations / acronyms and try to merge them to the cluster they belong to
    
    # Find named clusters that are all upper case, have ONLY one name cluster and at least two mentions
    abbreviation_clusters = [[x, grouped_name_clusters[x]] for x in grouped_name_clusters if grouped_name_clusters[x]!='Unnamed' and len(grouped_name_clusters[x].keys())==1 and len(grouped_name_clusters[x][list(grouped_name_clusters[x])[0]])>1 and (not re.findall(r'\W', list(grouped_name_clusters[x])[0])) and list(grouped_name_clusters[x])[0]==list(grouped_name_clusters[x])[0].upper()]

    # Find a grouped name cluster that has the same name as the abbreviation cluster

    # Strategy 1: Find the grouped name clusters that have the same name as the abbreviation cluster and group them all together
    to_merge = []
    for abbreviation_cluster_group, abbreviation_cluster in abbreviation_clusters:
        for grouped_name_cluster in grouped_name_clusters:
            if grouped_name_cluster != 'Unnamed':
                if list(grouped_name_clusters[grouped_name_cluster])[0] != list(abbreviation_cluster)[0]:
                    # # Method 1 : Use 'in'
                    # if any([list(abbreviation_cluster)[0] in x for x in list(grouped_name_clusters[grouped_name_cluster])]):
                    # Method 2 : Use regex
                    if any([re.findall(r'(?:^|\s|\(|\[){}(?:$|\s|\)|\])'.format(re.escape(list(abbreviation_cluster)[0])), x) for x in list(grouped_name_clusters[grouped_name_cluster])]):
                        # Add the cluster to merge
                        print('Merging {} and {}'.format(list(grouped_name_clusters[grouped_name_cluster]), list(abbreviation_cluster)[0]))
                        to_merge.append((grouped_name_cluster, abbreviation_cluster_group))
                    # Additional method : Check whether there is a grouped name cluster that has name clusters, which can abbreviate to the abbreviation cluster
                    elif any([''.join([y[0].upper() for y in re.split(r'\s+', x.strip()) if y])==list(abbreviation_cluster)[0] for x in list(grouped_name_clusters[grouped_name_cluster])]):
                        print('ADDITIONAL Merging {} and {}'.format(list(grouped_name_clusters[grouped_name_cluster]), list(abbreviation_cluster)[0]))
                        to_merge.append((grouped_name_cluster, abbreviation_cluster_group))
    
    # # Strategy 2: Find the grouped name clusters that have the same name as the abbreviation cluster and group the abbreviation cluster to the grouped name cluster that has the most matches
    # to_merge = []
    # for abbreviation_cluster_group, abbreviation_cluster in abbreviation_clusters:
    #     merge_counter = Counter()
    #     for grouped_name_cluster in grouped_name_clusters:
    #         if grouped_name_cluster != 'Unnamed':
    #             if list(grouped_name_clusters[grouped_name_cluster])[0] != list(abbreviation_cluster)[0]:
    #                 # # Method 1 : Use 'in'
    #                 # if any([list(abbreviation_cluster)[0] in x for x in list(grouped_name_clusters[grouped_name_cluster])]):
    #                 # Method 2 : Use regex
    #                 if any([re.findall(r'(?:^|\s|\(|\[){}(?:$|\s|\)|\])'.format(re.escape(list(abbreviation_cluster)[0])), x) for x in list(grouped_name_clusters[grouped_name_cluster])]):
    #                     # Add the cluster to merge
    #                     merge_counter[grouped_name_cluster] += 1
    #                 # Additional method : Check whether there is a grouped name cluster that has name clusters, which can abbreviate to the abbreviation cluster
    #                 elif any([''.join([y[0].upper() for y in re.split(r'\s+', x.strip()) if y])==list(abbreviation_cluster)[0] for x in list(grouped_name_clusters[grouped_name_cluster])]):
    #                     merge_counter[grouped_name_cluster] += 1
    #     if merge_counter:
    #         to_merge.append((merge_counter.most_common()[0][0], abbreviation_cluster_group))


    # Merge the clusters
    to_merge = cluster_tuples(to_merge)
    for cluster_group in to_merge:
        new_cluster_name = 'name_cluster_{}'.format('_'.join(['_'.join(x.split('_')[2:]) for x in cluster_group]))
        grouped_name_clusters[new_cluster_name] = {}
        for cluster in cluster_group:
            grouped_name_clusters[new_cluster_name].update(grouped_name_clusters[cluster])
            del grouped_name_clusters[cluster]

    return grouped_name_clusters


def make_qa_prompt(context: str, question: str) -> str:
    return f"""### Snippet: {context} ### Question: {question} ### Answer:"""


def classify(context: str, question: str, choices: tuple[str]) -> str:
    prompt = make_qa_prompt(context, question)
    json_data = {
            'text': prompt,
            'choices': choices,
            'top_k': 500,
            'gen_config': {
                'max_new_tokens': 256
            }
        }
    return classify_text(json_data['text'], json_data['choices'], json_data['top_k'], **json_data['gen_config'])


def answer(context: str, question: str) -> str:
    prompt = make_qa_prompt(context, question)
    json_data = {
            'text': prompt,
            'gen_config': {
                'max_new_tokens': 256,
            }
        }
    return infer_text(json_data['text'], **json_data['gen_config'])[0]


""" B Version """
map_cand_type = {
    'dataset': 'dataset',
    'software': 'software',
    'gaz_dataset': 'dataset',
    'gaz_method': 'software'
}

map_cand_type_plural = {
    'dataset': 'datasets',
    'software': 'software',
    'gaz_dataset': 'datasets',
    'gaz_method': 'software'
}

# Get gazeetteers from the synthetic and hybrid data
synthetic_external_gazetteers = dict()

with open(os.path.join(DATA_PATH, 'artifact_extraction_synthetic_data_v2.json'), 'r') as fin:
    synthetic_data = json.load(fin)

map_type_gaz = {
    'dataset': 'Dataset',
    'software': 'Method',
    'repository': 'Dataset',
    'method': 'Method'
}

synthetic_data_gaz = set([(re.search(r'<m>(.*)</m>', x['Snippet']).group(1), map_type_gaz[x['Type']]) for x in synthetic_data if x['Valid']=='Yes'])

# Add to the gazetteers
for gaz in synthetic_data_gaz:
    if gaz[0] in research_artifact_gazetteers and gaz[1] in research_artifact_gazetteers[gaz[0]]:
        continue
    if (gaz[0] in data_phrases and gaz[1]=='Dataset') or (gaz[0] in software_phrases and gaz[1]=='Method'):
        continue
    if gaz[0] not in synthetic_external_gazetteers:
        synthetic_external_gazetteers[gaz[0]] = [gaz[1]]
    if gaz[1] not in synthetic_external_gazetteers[gaz[0]]:
        synthetic_external_gazetteers[gaz[0]].append(gaz[1])

hybrid_external_gazetteers = dict()

with open(os.path.join(DATA_PATH, 'artifact_extraction_hybrid_data_v1.json'), 'r') as fin:
    hybrid_data = json.load(fin)

map_type_gaz = {
    'dataset': 'Dataset',
    'software': 'Method',
    'repository': 'Dataset',
    'method': 'Method'
}

hybrid_data_gaz = set([(re.search(r'<m>(.*)</m>', x['Snippet']).group(1), map_type_gaz[x['Type']]) for x in hybrid_data if x['Valid']=='Yes'])

# Add to the gazetteers
for gaz in hybrid_data_gaz:
    if gaz[0] in research_artifact_gazetteers and gaz[1] in research_artifact_gazetteers[gaz[0]]:
        continue
    if (gaz[0] in data_phrases and gaz[1]=='Dataset') or (gaz[0] in software_phrases and gaz[1]=='Method'):
        continue
    if gaz[0] not in hybrid_external_gazetteers:
        hybrid_external_gazetteers[gaz[0]] = [gaz[1]]
    if gaz[1] not in hybrid_external_gazetteers[gaz[0]]:
        hybrid_external_gazetteers[gaz[0]].append(gaz[1])


""" END B Version """

# FROM OLD CODE #

from urlextract import URLExtract

extractor = URLExtract()

cit_stopwords = ('leg', 'legend', 'tab', 'tabs', 'table', 'tables', 'fig', 'figs', 'figure', 'figures', 'note', 'notes', 'sup', 'suppl', 'supplementary', 'app', 'appx', 'append', 'appendix', 'ver', 'version', 'footnote', 'see', 'sec', 'section', 'cohort', 'ext', 'extended', 'additional')


def find_citations(text, stopwords=cit_stopwords):
    # TODO: some of the matches are duplicates
    matches_0 = re.findall(r'[[(][^\]\[)(]*?et al.*?(?:\]|\)|$)', text)
    matches_1 = re.findall(r'(\[\s*\d+\s*(?!\s*(?:\w|[^\w,\-\]])|\s*,\d+(?:[^\w,\-\]]))\s*?(?:\]|$))', text)
    matches_1b = re.findall(r'(\[\s*[A-Z](?:\w+\s+(?:\&\s+)?){1,}\d+\s*(?:\]|$))', text)
    matches_2 = re.findall(r'(\(\s*\d+\s*(?!\s*(?:\w|[^\w,\-\)])|\s*,\d+(?:[^\w,\-\)]))\s*?(?:\)|$))', text)
    matches_2b = re.findall(r'(\(\s*[A-Z](?:\w+\s+(?:\&\s+)?){1,}\d+\s*(?:\)|$))', text)
    # 1b, 2b are noisy, remove some stopwords
    matches_1b = [e for e in matches_1b if not any([e2.lower().strip() in stopwords for e2 in e.split()])]
    matches_2b = [e for e in matches_2b if not any([e2.lower().strip() in stopwords for e2 in e.split()])]
    return matches_0, matches_1, matches_1b, matches_2, matches_2b


# THIS IS INTENDED FOR USE IN THE LEFT PART OF THE CONCORDANCE (CHECK)
def has_citations(text):
    if 'et al' in text or re.search(r'\[\s*\d+', text) or re.search(r'\(\s*\d+\s*(?!\s*(?:\w|[^\w,\-\)])|\s*,\d+(?:[^\w,\-\)]))', text):
        return True
    return False


def filt_urls(url):
    is_valid = not re.match(r'^(?:\d+\.?)+$', url) and not re.match(r'^\d+\.\w+$', url)
    return is_valid


def find_urls(text, filter_gitlike=False, filter_noise=True):
    filters = ['github', 'gitlab', 'bitbucket', 'sourceforge', 'launchpad', 'cloud.google', 'aws.amazon',
               'phabricator', 'phacility', 'gogs', 'gitea', 'allura']  # research gate
    # urls = re.findall('http[s]?://(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\(\),]|(?:%[0-9a-fA-F][0-9a-fA-F]))+', text)
    urls = extractor.find_urls(text)
    # Solve a little bug that the extractor puts a special char in the end of the link
    urls = [e[:-1] if e[-1] in '.(){}[]<>,:!@$%^&*;' else e for e in urls]
    if filter_gitlike:
        try:
            filtered_urls = [e for e in urls if any([e2 in e.text for e2 in filters])]
        except:
            filtered_urls = [e for e in urls if any([e2 in e for e2 in filters])]
        return filtered_urls
    if filter_noise:
        return [e for e in urls if not (any([e[0].startswith(x) for x in ['i.e', 'et.al', 'e.g', 'al.']]) or any([e[0].endswith(x) for x in ['.et']])) and filt_urls(e)]
    return urls

# END OLD CODE #

# WITHOUT SCORE QUESTIONS
def extract_artifacts_candidate(cand_sent_id:str, cand_sent_text: str, cand_sent_paragraph: str, cand_sent_section: str, cand_type: str, cand_trigger: str, cand_trigger_off: tuple):
    # TODO: ADD BIGGER CONTEXT (PARAGRAPH, SECTION) FOR METADATA (MAYBE IT IS WISE TO DO IT IN A SECOND RUN WHEN NEEDED OR WITH ALL THE CONTEXT FROM "artifacts_cache"?)
    # TODO: IF WE FIND THAT IT IS THE SAME ARTIFACT AS ONE IN THE CACHE, WE CAN ASK ONLY THE QUESTIONS THAT HAVE NOT BEEN ANSWERED YET (THE N/A) (USING "candidates_metadata")
    # TODO: MAYBE WE CAN ASK WHETHER A PARAGRAPH OR SECTION IS WORTH CHECKING / IF IT HAS ANY RESEARCH ARTIFACTS (WILL SAVE SOME CREDITS)
    global artifacts_cache, candidates_metadata, thresholds

    result_dict = dict()

    # Background information so far
    background_info = ''

    # Create an annotated version of the sentence
    cand_sent_text_annot = cand_sent_text[:cand_trigger_off[0]] + '<m>' + cand_sent_text[cand_trigger_off[0]:cand_trigger_off[1]] + '</m>' + cand_sent_text[cand_trigger_off[1]:]

    # (WARNING: uncomment below to use "cand_sent_text_annot" instead of "cand_sent_text")
    cand_sent_text = cand_sent_text_annot

    # Check if valid artifact
    # TODO: CHANGE THIS WHEN YOU UPDATE THE MODEL
    # artifact_answer = classify(cand_sent_text, f'Is the artifact valid?', ('Yes', 'No'))
    artifact_answer = classify(cand_sent_text, f'Is there a valid {map_cand_type[cand_type]} defined in the <m> and </m> tags?', ('Yes', 'No'))
    
    result_dict['artifact_answer'] = artifact_answer
    artifact_answer_text = sorted(artifact_answer, key=lambda x:artifact_answer[x], reverse=True)[0]
    if artifact_answer_text == 'No' or artifact_answer['Yes'] < normalize_threshold(thresholds['artifact_answer'], len(artifact_answer)):
        return result_dict

    # Check name
    # Q: maybe we want to use the trigger word as the name if it comes from gazetteers? A: NO BECAUSE IT COULD BE A PART OF A NAME
    # Better solution is to include in the prompt that the name should contain the trigger word
    # TODO: CHANGE THIS WHEN YOU UPDATE THE MODEL
    # name_answer = answer(cand_sent_text, f'What is the name of artifact?')
    name_answer = answer(cand_sent_text, f'What is the name of the {map_cand_type[cand_type]} defined in the <m> and </m> tags?')

    
    # Check if the name is a keyword (excluding gazetteers)
    if name_answer.lower().strip() in data_phrases + software_phrases:
        # If it is, then the name is invalid
        name_answer = 'N/A'

    result_dict['name_answer'] = name_answer

    # Check license
    # TODO: CHANGE THIS WHEN YOU UPDATE THE MODEL
    # license_answer = answer(cand_sent_text, f'What is the license of the artifact?')
    license_answer = answer(cand_sent_text, f'What is the license of the {map_cand_type[cand_type]} defined in the <m> and </m> tags?')

    result_dict['license_answer'] = license_answer
    
    # Check version
    # TODO: CHANGE THIS WHEN YOU UPDATE THE MODEL
    # version_answer = answer(cand_sent_text, f'What is the version of the artifact?')
    version_answer = answer(cand_sent_text, f'What is the version of the {map_cand_type[cand_type]} defined in the <m> and </m> tags?')

    result_dict['version_answer'] = version_answer
    
    # Check url
    # TODO: CHANGE THIS WHEN YOU UPDATE THE MODEL
    # url_answer = answer(cand_sent_text, f'What is the URL of the artifact?')
    url_answer = answer(cand_sent_text, f'What is the URL of the {map_cand_type[cand_type]} defined in the <m> and </m> tags?')

    result_dict['url_answer'] = url_answer
    
    # Check ownership
    # TODO: CHANGE THIS WHEN YOU UPDATE THE MODEL
    # ownership_answer = classify(cand_sent_text, f'Is the artifact owned by the authors?', ('Yes', 'No'))
    ownership_answer = classify(cand_sent_text, f'Is the {map_cand_type[cand_type]} defined in the <m> and </m> tags introduced or created by the authors of the publication in the snippet above?', ('Yes', 'No'))

    ownership_answer_text = sorted(ownership_answer, key=lambda x:ownership_answer[x], reverse=True)[0]
    result_dict['ownership_answer'] = ownership_answer
    result_dict['ownership_answer_text'] = ownership_answer_text

    # Check reuse
    # TODO: CHANGE THIS WHEN YOU UPDATE THE MODEL
    # reuse_answer = classify(cand_sent_text, f'Is the artifact used by the authors?', ('Yes', 'No'))
    reuse_answer = classify(cand_sent_text, f'Is the {map_cand_type[cand_type]} defined in the <m> and </m> tags used or adopted by the authors of the publication in the snippet above?', ('Yes', 'No'))

    reuse_answer_text = sorted(reuse_answer, key=lambda x:reuse_answer[x], reverse=True)[0]
    result_dict['reuse_answer'] = reuse_answer
    result_dict['reuse_answer_text'] = reuse_answer_text

    return result_dict


def extract_artifacts_candidate_text_list(cand_sent_id:str, cand_sent_text: str, cand_type: str, cand_trigger: str, cand_trigger_off: tuple):
    # TODO: ADD BIGGER CONTEXT (PARAGRAPH, SECTION) FOR METADATA (MAYBE IT IS WISE TO DO IT IN A SECOND RUN WHEN NEEDED OR WITH ALL THE CONTEXT FROM "artifacts_cache"?)
    # TODO: IF WE FIND THAT IT IS THE SAME ARTIFACT AS ONE IN THE CACHE, WE CAN ASK ONLY THE QUESTIONS THAT HAVE NOT BEEN ANSWERED YET (THE N/A) (USING "candidates_metadata")
    # TODO: MAYBE WE CAN ASK WHETHER A PARAGRAPH OR SECTION IS WORTH CHECKING / IF IT HAS ANY RESEARCH ARTIFACTS (WILL SAVE SOME CREDITS)
    global artifacts_cache, candidates_metadata, thresholds

    result_dict = dict()

    # Background information so far
    background_info = ''

    # Create an annotated version of the sentence
    cand_sent_text_annot = cand_sent_text[:cand_trigger_off[0]] + '<m>' + cand_sent_text[cand_trigger_off[0]:cand_trigger_off[1]] + '</m>' + cand_sent_text[cand_trigger_off[1]:]

    # (WARNING: uncomment below to use "cand_sent_text_annot" instead of "cand_sent_text")
    cand_sent_text = cand_sent_text_annot

    # Check if valid artifact
    # TODO: CHANGE THIS WHEN YOU UPDATE THE MODEL
    # artifact_answer = classify(cand_sent_text, f'Is the artifact valid?', ('Yes', 'No'))
    artifact_answer = classify(cand_sent_text, f'Is there a valid {map_cand_type[cand_type]} defined in the <m> and </m> tags?', ('Yes', 'No'))
    
    result_dict['artifact_answer'] = artifact_answer
    artifact_answer_text = sorted(artifact_answer, key=lambda x:artifact_answer[x], reverse=True)[0]
    if artifact_answer_text == 'No' or artifact_answer['Yes'] < normalize_threshold(thresholds['artifact_answer'], len(artifact_answer)):
        return result_dict

    # Check name
    # Q: maybe we want to use the trigger word as the name if it comes from gazetteers? A: NO BECAUSE IT COULD BE A PART OF A NAME
    # Better solution is to include in the prompt that the name should contain the trigger word
    # TODO: CHANGE THIS WHEN YOU UPDATE THE MODEL
    # name_answer = answer(cand_sent_text, f'What is the name of artifact?')
    name_answer = answer(cand_sent_text, f'What is the name of the {map_cand_type[cand_type]} defined in the <m> and </m> tags?')

    
    # Check if the name is a keyword (excluding gazetteers)
    if name_answer.lower().strip() in data_phrases + software_phrases:
        # If it is, then the name is invalid
        name_answer = 'N/A'

    result_dict['name_answer'] = name_answer

    # Check license
    # TODO: CHANGE THIS WHEN YOU UPDATE THE MODEL
    # license_answer = answer(cand_sent_text, f'What is the license of the artifact?')
    license_answer = answer(cand_sent_text, f'What is the license of the {map_cand_type[cand_type]} defined in the <m> and </m> tags?')

    result_dict['license_answer'] = license_answer
    
    # Check version
    # TODO: CHANGE THIS WHEN YOU UPDATE THE MODEL
    # version_answer = answer(cand_sent_text, f'What is the version of the artifact?')
    version_answer = answer(cand_sent_text, f'What is the version of the {map_cand_type[cand_type]} defined in the <m> and </m> tags?')

    result_dict['version_answer'] = version_answer
    
    # Check url
    # TODO: CHANGE THIS WHEN YOU UPDATE THE MODEL
    # url_answer = answer(cand_sent_text, f'What is the URL of the artifact?')
    url_answer = answer(cand_sent_text, f'What is the URL of the {map_cand_type[cand_type]} defined in the <m> and </m> tags?')

    result_dict['url_answer'] = url_answer
    
    # Check ownership
    # TODO: CHANGE THIS WHEN YOU UPDATE THE MODEL
    # ownership_answer = classify(cand_sent_text, f'Is the artifact owned by the authors?', ('Yes', 'No'))
    ownership_answer = classify(cand_sent_text, f'Is the {map_cand_type[cand_type]} defined in the <m> and </m> tags introduced or created by the authors of the publication in the snippet above?', ('Yes', 'No'))

    ownership_answer_text = sorted(ownership_answer, key=lambda x:ownership_answer[x], reverse=True)[0]
    result_dict['ownership_answer'] = ownership_answer
    result_dict['ownership_answer_text'] = ownership_answer_text

    # Check reuse
    # TODO: CHANGE THIS WHEN YOU UPDATE THE MODEL
    # reuse_answer = classify(cand_sent_text, f'Is the artifact used by the authors?', ('Yes', 'No'))
    reuse_answer = classify(cand_sent_text, f'Is the {map_cand_type[cand_type]} defined in the <m> and </m> tags used or adopted by the authors of the publication in the snippet above?', ('Yes', 'No'))

    reuse_answer_text = sorted(reuse_answer, key=lambda x:reuse_answer[x], reverse=True)[0]
    result_dict['reuse_answer'] = reuse_answer
    result_dict['reuse_answer_text'] = reuse_answer_text

    return result_dict


def paragraph_section_check(paragraph_text: str):
    # This functions checks whether the paragraph contains any research artifacts worth checking
    # this could reduce the number of candidates to check, thus reducing the time, cost and noise
    result = answer(paragraph_text, f'List all artifacts in the above snippet.')
    return result


def check_paragraphs(pdf_metadata: dict, flat_candidates: list):
    # TODO: here we can also add the functionality to create new candidates from names of the artifacts found (!!! THIS COULD PARTLY SOLVE THE UNKNOWN NAMED ARTIFACTS PROBLEM !!!)
    paragraphs_counter = {k: Counter([(x[2], x[3]) for x in flat_candidates if x[0] == k]) for k in set([x[0] for x in flat_candidates])}
    paragraphs_to_check = sorted(set([i for s in [list(paragraphs_counter[x]) for x in paragraphs_counter if len(paragraphs_counter[x]) > 0] for i in s]), key=lambda y:(y[0], y[1]))
    paragraphs_to_analyze = dict()

    for paragraph_idx in tqdm(paragraphs_to_check):
        paragraph_text = ' '.join(pdf_metadata['sections'][paragraph_idx[0]][1][paragraph_idx[1]])
        # Check if the paragraph contains any research artifacts worth checking
        check_result = answer(paragraph_text, f'List all artifacts in the above snippet.')
        if check_result == 'N/A':
            continue
        else:
            cand_types_found = set([x.strip().split(':')[0].strip() for x in check_result.split('|')])
            for cand_type in cand_types_found:
                if cand_type not in paragraphs_to_analyze:
                    paragraphs_to_analyze[cand_type] = list()
                paragraphs_to_analyze[cand_type].append(paragraph_idx)

    return paragraphs_to_analyze


""" PIPELINE """

def normalize_threshold(threshold: float, n: int):
    if n <= 2:
        return threshold
    else:
        return threshold * (1 - (n-2)/n)  # Linearly decay the threshold with the number of options
        # return threshold * (1 - (n-2)/n)**0.5  # Square root decay the threshold with the number of options


def extract_research_artifacts_pipeline(pdf_metadata: dict, research_artifact_gazetteers: dict=research_artifact_gazetteers, data_phrases: list=data_phrases, software_phrases: list=software_phrases, filter_paragraphs=False, perform_deduplication=True, insert_fast_mode_gazetteers=False, dataset_gazetteers=None, verbose=True):
    global thresholds, artifacts_cache, candidates_metadata

    if dataset_gazetteers is None:
        external_gazetteers = dict()
    elif dataset_gazetteers == 'synthetic':
        external_gazetteers = synthetic_external_gazetteers
    elif dataset_gazetteers == 'hybrid':
        external_gazetteers = hybrid_external_gazetteers
    else:
        raise ValueError('Unknown dataset gazetteers')
    
    if insert_fast_mode_gazetteers:

        flat_text_list = []
        for i, section in enumerate(pdf_metadata['sections']):
            for j, paragraph in enumerate(section[1]):
                for k, sentence in enumerate(paragraph):
                    flat_text_list.append(sentence)
        new_gazetteers = []
        if verbose:
            t_flat_text_list = tqdm(flat_text_list, desc='Fast mode gazetteers')
        else:
            t_flat_text_list = flat_text_list
        for _text in t_flat_text_list:
            fast_mode_res = extract_research_artifacts_list_fast_mode([[_text]], research_artifact_gazetteers=research_artifact_gazetteers, data_phrases=data_phrases, software_phrases=software_phrases, dataset_gazetteers=dataset_gazetteers)
            for _res in fast_mode_res[0]['Results'].split('|'):
                if _res.strip() == 'N/A':
                    continue
                _res = _res.strip().split(':')
                if len(_res) != 2:
                    continue
                _type = _res[0].strip()
                _name = _res[1].strip()
                if _name.strip() == '':
                    continue
                if _type == 'software':
                    new_gazetteers.append([_name, 'Method'])
                elif _type == 'dataset':
                    new_gazetteers.append([_name, 'Dataset'])
                else:
                    print('Unknown type: {}'.format(_type))
                    continue
        # Add the new gazetteers to the existing gazetteers
        for new_gazetteer in new_gazetteers:
            if new_gazetteer[0] in research_artifact_gazetteers and new_gazetteer[1] in research_artifact_gazetteers[new_gazetteer[0]]:
                continue
            if (new_gazetteer[0] in data_phrases and new_gazetteer[1]=='Dataset') or (new_gazetteer[0] in software_phrases and new_gazetteer[1]=='Method'):
                continue
            if new_gazetteer[0] not in external_gazetteers:
                external_gazetteers[new_gazetteer[0]] = []
            if new_gazetteer[1] not in external_gazetteers[new_gazetteer[0]]:
                external_gazetteers[new_gazetteer[0]].append(new_gazetteer[1])
    
    candidates = detect_candidates(pdf_metadata, research_artifact_gazetteers, external_gazetteers, data_phrases, software_phrases)

    flat_candidates = sorted([i2 for s2 in [[i for s in [[[ct, k, y[0], y[1], y[2], y[3]] for y in candidates[ct][k]] for k in candidates[ct]] for i in s] for ct in candidates] for i2 in s2], key=lambda z:(z[2], z[3], z[4], z[5][0], z[5][1]))

    # Initialize/Reset the cache of artifacts for this pdf (maybe we could create persistent PDF identifiers)
    #  this cache will store the artifacts that have been found in the pdf, along with their snippets
    #  (maybe) this cache will also store some of the metadata (e.g. name, ownership, reuse, etc)
    artifacts_cache = dict()

    # Initialize/Reset the candidates metadata for this pdf
    candidates_metadata = dict()

    # Find coreference chains for the candidates
    candidate_annotated_sentences = [[x[0], x[1], x[2], x[3], x[4], x[5], x_i, pdf_metadata['sections'][x[2]][1][x[3]][x[4]][:x[5][0]] + '<m>' + pdf_metadata['sections'][x[2]][1][x[3]][x[4]][x[5][0]:x[5][1]] + '</m>' + pdf_metadata['sections'][x[2]][1][x[3]][x[4]][x[5][1]:]] for x_i, x in enumerate(flat_candidates)]
    
    # Group the sentences of the same paragraphs
    candidate_annotated_paragraphs = dict()
    for x in candidate_annotated_sentences:
        if (x[2], x[3]) not in candidate_annotated_paragraphs:
            candidate_annotated_paragraphs[(x[2], x[3])] = dict()
        if map_cand_type[x[0]] not in candidate_annotated_paragraphs[(x[2], x[3])]:
            candidate_annotated_paragraphs[(x[2], x[3])][map_cand_type[x[0]]] = list()
        candidate_annotated_paragraphs[(x[2], x[3])][map_cand_type[x[0]]].append(x)

    if filter_paragraphs:
        # Check the paragraph/sections if they contain any research artifacts worth checking
        valid_paragraph_indices = check_paragraphs(pdf_metadata, flat_candidates)

    # Filter the candidates based on the valid paragraph indices
    for cand in flat_candidates:
        if (not filter_paragraphs) or (map_cand_type[cand[0]] in valid_paragraph_indices and ((cand[2], cand[3]) in valid_paragraph_indices[map_cand_type[cand[0]]])):
            cand.append('analyze')
        else:
            cand.append('skip')
    
    # discarded_candidates = [[x, ' '.join(pdf_metadata['sections'][x[2]][1][x[3]])] for x in flat_candidates if (x[2], x[3]) not in valid_paragraph_indices[x[0]]]

    if verbose:
        t_flat_candidates = tqdm(enumerate(flat_candidates), total=len(flat_candidates))
    else:
        t_flat_candidates = enumerate(flat_candidates)
    for cand_sent_idx, cand_sent in t_flat_candidates:
        # Candidate sentence id
        cand_sent_id = f'C{cand_sent_idx}'
        # Get candidate sentence text
        cand_sent_text = pdf_metadata['sections'][cand_sent[2]][1][cand_sent[3]][cand_sent[4]]
        # Get candidate sentence paragraph
        cand_sent_paragraph = ' '.join(pdf_metadata['sections'][cand_sent[2]][1][cand_sent[3]])
        # Get candidate sentence section
        cand_sent_section = ' '.join([i for s in pdf_metadata['sections'][cand_sent[2]][1] for i in s])

        # # This is a dictionary that contains all information about the candidates
        # cand_sent.append(dict())

        # Find citations
        cand_sent_citations = find_citations(cand_sent_text)

        # Find urls
        cand_sent_urls = find_urls(cand_sent_text)

        snippet_start = len(' '.join(pdf_metadata['sections'][cand_sent[2]][1][cand_sent[3]][:cand_sent[4]]) + ' ') if pdf_metadata['sections'][cand_sent[2]][1][cand_sent[3]][:cand_sent[4]] else 0
        snippet_end = len(' '.join(pdf_metadata['sections'][cand_sent[2]][1][cand_sent[3]][:cand_sent[4]])) + len(cand_sent_text) + 1
        if len(pdf_metadata['sections'][cand_sent[2]][1][cand_sent[3]])-1!=cand_sent[4]:
            snippet_end -= 1

        # Skip the candidates that are not in the valid paragraphs
        if cand_sent[-1] == 'skip':
            candidates_metadata[cand_sent_id] = {
                'type': cand_sent[0],
                'indices': (cand_sent[2], cand_sent[3], cand_sent[4]),
                'trigger': cand_sent[1],
                'trigger_offset': cand_sent[5],
                'snippet': cand_sent_text,
                'snippet_offset': (snippet_start, snippet_end),
                'paragraph': cand_sent_paragraph,
                'paragraph_offset': (len(' '.join([i for s in pdf_metadata['sections'][cand_sent[2]][1][:cand_sent[3]] for i in s]) + ' '), len(' '.join([i for s in pdf_metadata['sections'][cand_sent[2]][1][:cand_sent[3]] for i in s]) + ' ') + len(cand_sent_paragraph)),
                'section': cand_sent_section,
                'section_title': pdf_metadata['sections'][cand_sent[2]][0],
                'citations': cand_sent_citations,
                'urls': cand_sent_urls,
                'results': None,
                'skipped': True
            }
        else:
            # Add the results to the metadata dictionary
            candidates_metadata[cand_sent_id] = {
                'type': cand_sent[0],
                'indices': (cand_sent[2], cand_sent[3], cand_sent[4]),
                'trigger': cand_sent[1],
                'trigger_offset': cand_sent[5],
                'snippet': cand_sent_text,
                'snippet_offset': (snippet_start, snippet_end),
                'paragraph': cand_sent_paragraph,
                'paragraph_offset': (len(' '.join([i for s in pdf_metadata['sections'][cand_sent[2]][1][:cand_sent[3]] for i in s]) + ' '), len(' '.join([i for s in pdf_metadata['sections'][cand_sent[2]][1][:cand_sent[3]] for i in s]) + ' ') + len(cand_sent_paragraph)),
                'section': cand_sent_section,
                'section_title': pdf_metadata['sections'][cand_sent[2]][0],
                'citations': cand_sent_citations,
                'urls': cand_sent_urls,
                'results': extract_artifacts_candidate(cand_sent_id, cand_sent_text, cand_sent_paragraph, cand_sent_section, cand_sent[0], cand_sent[1], cand_sent[5]),
                'skipped': False
            }

    if perform_deduplication:
        # Deduplicate the mentions per type
        all_types = sorted(set(map_cand_type.values()))
        grouped_clusters = dict()
        for typ in all_types:
            mentions = [[x, candidates_metadata[x], candidates_metadata[x]['paragraph'][:candidates_metadata[x]['snippet_offset'][0]] + candidates_metadata[x]['snippet'][:candidates_metadata[x]['trigger_offset'][0]] + '<m>' + candidates_metadata[x]['snippet'][candidates_metadata[x]['trigger_offset'][0]:candidates_metadata[x]['trigger_offset'][1]] + '</m>' + candidates_metadata[x]['snippet'][candidates_metadata[x]['trigger_offset'][1]:] + candidates_metadata[x]['paragraph'][candidates_metadata[x]['snippet_offset'][1]:]] for x in candidates_metadata if not candidates_metadata[x]['skipped'] and 'name_answer' in candidates_metadata[x]['results'] and map_cand_type[candidates_metadata[x]['type']]==typ]
            grouped_clusters[typ] = deduplicate_mentions_dedup(mentions)

        return candidates_metadata, grouped_clusters
    else:
        return candidates_metadata


def create_ordered_set(items):
    return list(OrderedDict.fromkeys(items))


def extract_research_artifacts_fast_mode(pdf_metadata, research_artifact_gazetteers: dict=research_artifact_gazetteers, data_phrases: list=data_phrases, software_phrases: list=software_phrases, dataset_gazetteers=None, verbose=True):
    if dataset_gazetteers is None:
        external_gazetteers = dict()
    elif dataset_gazetteers == 'synthetic':
        external_gazetteers = synthetic_external_gazetteers
    elif dataset_gazetteers == 'hybrid':
        external_gazetteers = hybrid_external_gazetteers
    else:
        raise ValueError('Unknown dataset gazetteers')
    
    candidates = detect_candidates(pdf_metadata, research_artifact_gazetteers, external_gazetteers, data_phrases, software_phrases)

    flat_candidates = sorted([i2 for s2 in [[i for s in [[[ct, k, y[0], y[1], y[2], y[3]] for y in candidates[ct][k]] for k in candidates[ct]] for i in s] for ct in candidates] for i2 in s2], key=lambda z:(z[2], z[3], z[4], z[5][0], z[5][1]))

    paragraph_results = list()
    t_pdf_metadata_sections = tqdm(pdf_metadata['sections']) if verbose else pdf_metadata['sections']
    for i, section in enumerate(t_pdf_metadata_sections):
        for j, paragraph in enumerate(section[1]):
            paragraph_text = ' '.join(paragraph)
            paragraph_result = answer(paragraph_text, f'List all artifacts in the above snippet.')
            paragraph_results.append({
                'Paragraph': paragraph_text,
                'Candidates': ' | '.join(create_ordered_set(['[{}] : {}'.format(map_cand_type[x[0]], x[1]) for x in flat_candidates if x[2]==i and x[3]==j])),
                'Results': paragraph_result
            })

    return paragraph_results


def extract_research_artifacts_list(text_list, research_artifact_gazetteers=research_artifact_gazetteers, data_phrases=data_phrases, software_phrases=software_phrases, perform_deduplication=True, insert_fast_mode_gazetteers=False, dataset_gazetteers=None,verbose=True):
    # TODO: UPDATE THIS WHEN ADDING MORE RESEARCH ARTIFACTS AND/OR GAZETTEERS
    if dataset_gazetteers is None:
        external_gazetteers = dict()
    elif dataset_gazetteers == 'synthetic':
        external_gazetteers = synthetic_external_gazetteers
    elif dataset_gazetteers == 'hybrid':
        external_gazetteers = hybrid_external_gazetteers
    else:
        raise ValueError('Unknown dataset gazetteers')
    if insert_fast_mode_gazetteers:
        new_gazetteers = []
        flat_text_list = [i for s in text_list for i in s]
        for _text in flat_text_list:
            fast_mode_res = extract_research_artifacts_list_fast_mode([[_text]], research_artifact_gazetteers=research_artifact_gazetteers, data_phrases=data_phrases, software_phrases=software_phrases, dataset_gazetteers=dataset_gazetteers)
            for _res in fast_mode_res[0]['Results'].split('|'):
                if _res.strip() == 'N/A':
                    continue
                _res = _res.strip().split(':')
                if len(_res) != 2:
                    continue
                _type = _res[0].strip()
                _name = _res[1].strip()
                if _name.strip() == '':
                    continue
                if _type == 'software':
                    new_gazetteers.append([_name, 'Method'])
                elif _type == 'dataset':
                    new_gazetteers.append([_name, 'Dataset'])
                else:
                    print('Unknown type: {}'.format(_type))
                    continue
        # Add the new gazetteers to the existing gazetteers
        for new_gazetteer in new_gazetteers:
            if new_gazetteer[0] in research_artifact_gazetteers and new_gazetteer[1] in research_artifact_gazetteers[new_gazetteer[0]]:
                continue
            if (new_gazetteer[0] in data_phrases and new_gazetteer[1]=='Dataset') or (new_gazetteer[0] in software_phrases and new_gazetteer[1]=='Method'):
                continue
            if new_gazetteer[0] not in external_gazetteers:
                external_gazetteers[new_gazetteer[0]] = []
            if new_gazetteer[1] not in external_gazetteers[new_gazetteer[0]]:
                external_gazetteers[new_gazetteer[0]].append(new_gazetteer[1])

    candidates = detect_candidates_list(text_list, research_artifact_gazetteers, external_gazetteers, data_phrases, software_phrases)

    flat_candidates = sorted([i2 for s2 in [[i for s in [[[ct, k, y[0], y[1], y[2]] for y in candidates[ct][k]] for k in candidates[ct]] for i in s] for ct in candidates] for i2 in s2], key=lambda z:(z[2], z[3], z[4]))
    
    # Initialize/Reset the candidates metadata for this pdf
    candidates_metadata = dict()

    if verbose:
        t_flat_candidates = tqdm(enumerate(flat_candidates), total=len(flat_candidates))
    else:
        t_flat_candidates = enumerate(flat_candidates)
    for cand_sent_idx, cand_sent in t_flat_candidates:
        # Candidate sentence id
        cand_sent_id = f'C{cand_sent_idx}'
        # Get candidate sentence text
        cand_sent_text = text_list[cand_sent[2]][cand_sent[3]]
        # Get candidate sentence paragraph
        cand_sent_paragraph = ' '.join(text_list[cand_sent[2]])
        # Find citations
        cand_sent_citations = find_citations(cand_sent_text)

        # Find snippet offsets
        # TODO: MAYBE THIS (FIXED) SHOULD BE COPIED TO THE ANALYZE IT MAIN CODE??
        snippet_start = len(' '.join(text_list[cand_sent[2]][:cand_sent[3]]) + ' ') if text_list[cand_sent[2]][:cand_sent[3]] else 0
        snippet_end = len(' '.join(text_list[cand_sent[2]][:cand_sent[3]])) + len(cand_sent_text) + 1
        if len(text_list[cand_sent[2]])-1!=cand_sent[3]:
            snippet_end -= 1

        # Add the results to the metadata dictionary
        candidates_metadata[cand_sent_id] = {
            'type': cand_sent[0],
            'indices': (cand_sent[2], cand_sent[3]),
            'trigger': cand_sent[1],
            'trigger_offset': cand_sent[4],
            'snippet': cand_sent_text,
            'snippet_offset': (snippet_start, snippet_end),  # TODO: MAYBE THIS LINE (FIXED) SHOULD BE COPIED TO THE ANALYZE IT MAIN CODE??
            'paragraph': cand_sent_paragraph,
            'citations': cand_sent_citations,
            'results': extract_artifacts_candidate_text_list(cand_sent_id, cand_sent_text, cand_sent[0], cand_sent[1], cand_sent[4]),
        }
    
    if perform_deduplication:
        # Deduplicate the mentions per type
        all_types = sorted(set(map_cand_type.values()))
        grouped_clusters = dict()
        for typ in all_types:
            mentions = [[x, candidates_metadata[x], candidates_metadata[x]['paragraph'][:candidates_metadata[x]['snippet_offset'][0]] + candidates_metadata[x]['snippet'][:candidates_metadata[x]['trigger_offset'][0]] + '<m>' + candidates_metadata[x]['snippet'][candidates_metadata[x]['trigger_offset'][0]:candidates_metadata[x]['trigger_offset'][1]] + '</m>' + candidates_metadata[x]['snippet'][candidates_metadata[x]['trigger_offset'][1]:] + candidates_metadata[x]['paragraph'][candidates_metadata[x]['snippet_offset'][1]:]] for x in candidates_metadata if 'name_answer' in candidates_metadata[x]['results'] and map_cand_type[candidates_metadata[x]['type']]==typ]
            grouped_clusters[typ] = deduplicate_mentions_dedup(mentions)
    
        return candidates_metadata, grouped_clusters
    else:
        return candidates_metadata
    

def extract_research_artifacts_list_fast_mode(text_list, research_artifact_gazetteers: dict=research_artifact_gazetteers, data_phrases: list=data_phrases, software_phrases: list=software_phrases, dataset_gazetteers=None, verbose=False):
    if dataset_gazetteers is None:
        external_gazetteers = dict()
    elif dataset_gazetteers == 'synthetic':
        external_gazetteers = synthetic_external_gazetteers
    elif dataset_gazetteers == 'hybrid':
        external_gazetteers = hybrid_external_gazetteers
    else:
        raise ValueError('Unknown dataset gazetteers')
    candidates = detect_candidates_list(text_list, research_artifact_gazetteers, external_gazetteers, data_phrases, software_phrases)

    flat_candidates = sorted([i2 for s2 in [[i for s in [[[ct, k, y[0], y[1], y[2]] for y in candidates[ct][k]] for k in candidates[ct]] for i in s] for ct in candidates] for i2 in s2], key=lambda z:(z[2], z[3], z[4]))

    paragraph_results = list()
    if verbose:
        t_text_list = tqdm(enumerate(text_list), total=len(text_list))
    else:
        t_text_list = enumerate(text_list)
    for i, paragraph in t_text_list:
        paragraph_text = ' '.join(paragraph)
        paragraph_result = answer(paragraph_text, f'List all artifacts in the above snippet.')
        paragraph_results.append({
            'Paragraph': paragraph_text,
            'Candidates': ' | '.join(create_ordered_set(['[{}] : {}'.format(map_cand_type[x[0]], x[1]) for x in flat_candidates if x[2]==i])),
            'Results': paragraph_result
        })

    return paragraph_results

"""

RE-EVALUATE FUNCTION

"""

import pandas as pd
from openpyxl.styles import Font
from openpyxl import load_workbook


def is_valid_section_title(section_title):
    if not section_title:
        return True
    elif not re.findall(r'\brelated\b', section_title.lower(), re.IGNORECASE) and not re.findall(r'\bbackground\b', section_title.lower(), re.IGNORECASE):
        return True
    else:
        return False
    

def reevaluate(file_path, new_thresholds=None):

    if new_thresholds is None:
        new_thresholds = {
            'artifact_answer': 0.8,
            'ownership_answer': 0.8,
            'reuse_answer': 0.8,
    }


    output_file = file_path.replace('.json', '_reevaluated.xlsx')

    with open(file_path, 'r') as f:
        output_json = json.load(f)

    # Create a dataframe with the information about the publication: 'Title', 'Abstract', 'Identifiers', 'Claim', 'Claimer', 'Topic'
    publication_info = {
        'Identifiers': '\n'.join(['{}: {}'.format(x[0], x[1]) for x in output_json['pdf_metadata']['identifiers']]),
        'Title': output_json['pdf_metadata']['title'],
        'Authors': '\n'.join([(x['forename'] + ' ' if x['forename'] else '') + (x['surname'] if x['surname'] else '') for x in output_json['pdf_metadata']['bibl_references'][0]['authors']]),
        'Abstract': ' '.join([i for s in output_json['pdf_metadata']['abstract'] for i in s]),
        'Sections': '\n'.join([x[0] for x in output_json['pdf_metadata']['sections'] if x[0]]),
        'Claim': output_json['claim'] if 'claim' in output_json else 'Skipped',
        'Claimer': output_json['claimer'] if 'claimer' in output_json else 'Skipped',
        'Topic': output_json['topic'] if 'topic' in output_json else 'Skipped',
        'FOS Pred 1': 'L1: {}\nL2: {}\nL3: {}\nL4: {}\nL5: {}\nL6: {}'.format(
            output_json['fos_results']['fos_result'][0]['L1'], 
            output_json['fos_results']['fos_result'][0]['L2'], 
            output_json['fos_results']['fos_result'][0]['L3'], 
            output_json['fos_results']['fos_result'][0]['L4'] if 'L4' in output_json['fos_results']['fos_result'][0] else 'N/A',
            output_json['fos_results']['fos_result'][0]['L5'] if 'L5' in output_json['fos_results']['fos_result'][0] else 'N/A',
            output_json['fos_results']['fos_result'][0]['L6'] if 'L6' in output_json['fos_results']['fos_result'][0] else 'N/A'
            ) if 'fos_results' in output_json else 'Skipped',
        'FOS Pred 2': 'L1: {}\nL2: {}\nL3: {}\nL4: {}\nL5: {}\nL6: {}'.format(
            output_json['fos_results']['fos_result'][1]['L1'], 
            output_json['fos_results']['fos_result'][1]['L2'], 
            output_json['fos_results']['fos_result'][1]['L3'], 
            output_json['fos_results']['fos_result'][1]['L4'] if 'L4' in output_json['fos_results']['fos_result'][1] else 'N/A',
            output_json['fos_results']['fos_result'][1]['L5'] if 'L5' in output_json['fos_results']['fos_result'][1] else 'N/A',
            output_json['fos_results']['fos_result'][1]['L6'] if 'L6' in output_json['fos_results']['fos_result'][1] else 'N/A'
            ) if 'fos_results' in output_json else 'Skipped',
        'FOS Pred 3': 'L1: {}\nL2: {}\nL3: {}\nL4: {}\nL5: {}\nL6: {}'.format(
            output_json['fos_results']['fos_result'][2]['L1'], 
            output_json['fos_results']['fos_result'][2]['L2'], 
            output_json['fos_results']['fos_result'][2]['L3'], 
            output_json['fos_results']['fos_result'][2]['L4'] if 'L4' in output_json['fos_results']['fos_result'][2] else 'N/A',
            output_json['fos_results']['fos_result'][2]['L5'] if 'L5' in output_json['fos_results']['fos_result'][2] else 'N/A',
            output_json['fos_results']['fos_result'][2]['L6'] if 'L6' in output_json['fos_results']['fos_result'][2] else 'N/A'
            ) if 'fos_results' in output_json else 'Skipped',
    }
    publication_info_df = pd.DataFrame(publication_info, index=['']).T

    # Create a dataframe with the information about the bibliographic references (citations) and the citances
    citations = []
    for bibl_ref in output_json['pdf_metadata']['bibl_references'][1:]:
        # Citations
        citations.append({
            'Id': bibl_ref['id'] if bibl_ref['id'] else '',
            'Doi': bibl_ref['doi'] if bibl_ref['doi'] else '',
            'Title': bibl_ref['title'] if bibl_ref['title'] else '',
            'Authors': '\n'.join([(x['forename'] + ' ' if x['forename'] else '')+(x['surname'] if x['surname'] else '') for x in bibl_ref['authors']]) if bibl_ref['authors'] else '',
        })
    citations = sorted(citations, key=lambda x: int(x['Id'].replace('b', '')))
    # Sort citations by id
    citations_df = pd.DataFrame(citations)

    citances = []
    for citance in output_json['pdf_metadata']['citances']:
        # Citances
        for ref in citance['refs']:
            citances.append({
                'Target': ref['target'] if ref['target'] else '',
                'Type': ref['type'] if ref['type'] else '',
                'Indices': (citance['sec_idx'], citance['par_idx'], citance['s_idx']),
                'CitText': ref['text'] if ref['text'] else '',
                'Intent': (ref['intent'] if ref['intent'] else '') if 'intent' in ref else 'Skipped',
                'Polarity': (ref['polarity'] if ref['polarity'] else '') if 'polarity' in ref else 'Skipped',
                'Semantics': (ref['semantics'] if ref['semantics'] else '') if 'semantics' in ref else 'Skipped',
                'Intent Score': (ref['intent_probs'][ref['intent']] if ref['intent_probs'] else '') if 'intent' in ref else 'Skipped',
                'Polarity Score': (ref['polarity_probs'][ref['polarity']] if ref['polarity_probs'] else '') if 'polarity' in ref else 'Skipped',
                'Semantics Score': (ref['semantics_probs'][ref['semantics']] if ref['semantics_probs'] else '') if 'semantics' in ref else 'Skipped',
                'Sentence': citance['sentence'] if citance['sentence'] else '',
            })
    # Sort citances by target (convert #b1 to 1, #b2 to 2, etc. and sort) (if '_' is present, then it is a reference to a table or figure, so first sort by the text and then by the number after the '_')
    citances = sorted(citances, key=lambda x: (x['Target'].split('_')[0], int(x['Target'].split('_')[1])) if '_' in x['Target'] else ('', int(x['Target'].replace('#b', '')) if x['Target']!='' else -1))
    citances_df = pd.DataFrame(citances)

    """ NEW STUFF / THRESHOLDS """
    from copy import deepcopy

    output_mentions_list = sorted([i3 for s3 in [i2 for s2 in [i for s in [[[output_json['research_artifacts']['grouped_clusters'][k][k2][k3] for k3 in output_json['research_artifacts']['grouped_clusters'][k][k2]] for k2 in output_json['research_artifacts']['grouped_clusters'][k]] for k in output_json['research_artifacts']['grouped_clusters']] for i in s] for i2 in s2] for i3 in s3], key=lambda y:int(y[0][1:]))
    
    # Apply the new thresholds to the mentions of the research artifacts
    new_mentions_list = deepcopy(output_mentions_list)
    to_remove = []
    for mention in new_mentions_list:
        if (mention[1]['results']['artifact_answer']['Yes'] >= new_thresholds['artifact_answer']):
            if (mention[1]['results']['ownership_answer']['Yes'] >= new_thresholds['ownership_answer']):
                mention[1]['results']['ownership_final_answer'] = 'Yes'
            else:
                mention[1]['results']['ownership_final_answer'] = 'No'

            if (mention[1]['results']['reuse_answer']['Yes'] >= new_thresholds['reuse_answer']):
                mention[1]['results']['reuse_final_answer'] = 'Yes'
            else:
                mention[1]['results']['reuse_final_answer'] = 'No'
        else:
            to_remove.append(mention)
    
    # Remove the mentions that do not meet the new thresholds
    for mention in to_remove:
        new_mentions_list.remove(mention)
    
    # Cluster again the research artifacts
    mention_types = set([map_cand_type[x[1]['type']] if x[1]['type'] in map_cand_type else x[1]['type'] for x in new_mentions_list])

    research_artifacts = {}
    for m_type in mention_types:
        m_type_mentions = [x for x in new_mentions_list if (map_cand_type[x[1]['type']] if x[1]['type'] in map_cand_type else x[1]['type']) == m_type]
        research_artifacts[m_type] = deduplicate_mentions_dedup(m_type_mentions)
    
    # Assign the new clusters to the output
    output_json['research_artifacts']['grouped_clusters'] = research_artifacts

    """ NEW STUFF / THRESHOLDS -- END """

    # Better "visualize" the research artifacts (print and add to excel)
    columns = ['RA Cluster', 'Research Artifact', 'Type', 'Research Artifact Score', 'Owned', 'Owned Percentage', 'Owned Score', 'Reused', 'Reused Percentage', 'Reused Score', 'Licenses', 'Versions', 'URLs', 'Citations', 'Mentions Count']
    df = pd.DataFrame(columns=columns)

    mentions_columns = ['Mention ID', 'RA Cluster', 'Research Artifact', 'Type', 'Research Artifact Score', 'Owned', 'Owned Score', 'Reused', 'Reused Score', 'License', 'Version', 'URLs', 'Citations', 'Section', 'Indices', 'Trigger', 'Mention']
    mentions_df = pd.DataFrame(columns=mentions_columns)

    for ra_type in output_json['research_artifacts']['grouped_clusters']:
        for ra_cluster in output_json['research_artifacts']['grouped_clusters'][ra_type]:
            for ra_name in output_json['research_artifacts']['grouped_clusters'][ra_type][ra_cluster]:
                # Find research artifact scores
                ra_scores = [x[1]['results']['artifact_answer']['Yes'] for x in output_json['research_artifacts']['grouped_clusters'][ra_type][ra_cluster][ra_name]]
                mean_ra_score = sum(ra_scores) / len(ra_scores)

                """ NEW STUFF / OWNERSHIP """
                k = 1
                valid_onwership_mentions = sorted([x for x in output_json['research_artifacts']['grouped_clusters'][ra_type][ra_cluster][ra_name] if is_valid_section_title(x[1]['section_title'])], key=lambda y: (y[1]['indices'][0], y[1]['indices'][1], y[1]['indices'][2], int(y[0][1:])))[:k]
                ownership_counter = Counter([x[1]['results']['ownership_final_answer'] for x in valid_onwership_mentions])
                """ NEW STUFF / OWNERSHIP -- END """

                # reuse_counter = Counter([sorted(x[1]['results']['reuse_answer'], key=lambda y: x[1]['results']['reuse_answer'][y], reverse=True)[0] for x in output_json['research_artifacts']['grouped_clusters'][ra_type][ra_cluster][ra_name]])  # TODO: THIS WAS A BUG IN DTH -- PLEASE INFER AGAIN
                reuse_counter = Counter([x[1]['results']['reuse_final_answer'] for x in output_json['research_artifacts']['grouped_clusters'][ra_type][ra_cluster][ra_name]])

                owned = 'Yes' if 'Yes' in ownership_counter else 'No'
                owned_percentage = ownership_counter['Yes'] / len(output_json['research_artifacts']['grouped_clusters'][ra_type][ra_cluster][ra_name]) * 100
                reused = 'Yes' if 'Yes' in reuse_counter else 'No'
                reused_percentage = reuse_counter['Yes'] / len(output_json['research_artifacts']['grouped_clusters'][ra_type][ra_cluster][ra_name]) * 100
                owned_scores = [x[1]['results']['ownership_answer']['Yes'] for x in output_json['research_artifacts']['grouped_clusters'][ra_type][ra_cluster][ra_name] if sorted(x[1]['results']['ownership_answer'], key=lambda y: x[1]['results']['ownership_answer'][y], reverse=True)[0]=='Yes']
                mean_owned_score = sum(owned_scores) / len(owned_scores) if owned_scores else 0
                reused_scores = [x[1]['results']['reuse_answer']['Yes'] for x in output_json['research_artifacts']['grouped_clusters'][ra_type][ra_cluster][ra_name] if sorted(x[1]['results']['reuse_answer'], key=lambda y: x[1]['results']['reuse_answer'][y], reverse=True)[0]=='Yes']
                mean_reused_score = sum(reused_scores) / len(reused_scores) if reused_scores else 0

                # Find licenses
                license_counter = Counter([x[1]['results']['license_answer'] for x in output_json['research_artifacts']['grouped_clusters'][ra_type][ra_cluster][ra_name] if x[1]['results']['license_answer']!='N/A'])
                # Find the percentage of each license
                all_licenses_count = sum(license_counter.values())
                all_licenses = [[license, license_counter[license] / all_licenses_count * 100] for license in license_counter]

                # Find versions
                version_counter = Counter([x[1]['results']['version_answer'] for x in output_json['research_artifacts']['grouped_clusters'][ra_type][ra_cluster][ra_name] if x[1]['results']['version_answer']!='N/A'])
                # Find the percentage of each version
                all_versions_count = sum(version_counter.values())
                all_versions = [[version, version_counter[version] / all_versions_count * 100] for version in version_counter]

                # Find URLs
                urls = Counter([i for s in [x[1]['urls'] for x in output_json['research_artifacts']['grouped_clusters'][ra_type][ra_cluster][ra_name] if x[1]['urls']] for i in s])
                # Find the percentage of each URL
                all_urls_count = sum(urls.values())
                all_urls = [[url, urls[url] / all_urls_count * 100] for url in urls]

                # Find Citations
                citations = Counter([i2 for s2 in [i for s in [x[1]['citations'] for x in output_json['research_artifacts']['grouped_clusters'][ra_type][ra_cluster][ra_name] if x[1]['citations']] for i in s] for i2 in s2])
                # Find the percentage of each citation
                all_citations_count = sum(citations.values())
                all_citations = [[citation, citations[citation] / all_citations_count * 100] for citation in citations]

                # print(f'Research Artifact: {ra_name} ({ra_type})')
                # print(f'Owned: {owned} ({owned_percentage}%)')
                # print(f'Reused: {reused} ({reused_percentage}%)')
                # print(f'Licenses: [{", ".join([f"{license} ({percentage}%)" for license, percentage in all_licenses])}]')
                # print(f'Versions: [{", ".join([f"{version} ({percentage}%)" for version, percentage in all_versions])}]')
                # print(f'URLs: [{", ".join([f"{url} ({percentage}%)" for url, percentage in all_urls])}]')
                # print(f'Citations: [{", ".join([f"{citation} ({percentage}%)" for citation, percentage in all_citations])}]')
                # print()

                # Find mentions
                mentions = [
                    [
                        x[1]['snippet'][:x[1]['trigger_offset'][0]] + '<m>' + x[1]['snippet'][x[1]['trigger_offset'][0]:x[1]['trigger_offset'][1]] + '</m>' + x[1]['snippet'][x[1]['trigger_offset'][1]:], 
                        x[0],
                        # map_cand_type[x[1]['type']],
                        x[1]['type'],
                        x[1]['results']['artifact_answer']['Yes'],
                        'Yes' if sorted(x[1]['results']['ownership_answer'], key=lambda y: x[1]['results']['ownership_answer'][y], reverse=True)[0]=='Yes' else 'No',
                        x[1]['results']['ownership_answer']['Yes'],
                        'Yes' if sorted(x[1]['results']['reuse_answer'], key=lambda y: x[1]['results']['reuse_answer'][y], reverse=True)[0]=='Yes' else 'No',
                        x[1]['results']['reuse_answer']['Yes'],
                        '' if x[1]['results']['license_answer']=='N/A' else x[1]['results']['license_answer'],
                        '' if x[1]['results']['version_answer']=='N/A' else x[1]['results']['version_answer'],
                        "\n".join(x[1]['urls']),
                        "\n".join(sorted(set([i for s in x[1]['citations'] for i in s]))),
                        x[1]['section_title'],
                        tuple(x[1]['indices']),
                        x[1]['trigger'],
                    ] for x in output_json['research_artifacts']['grouped_clusters'][ra_type][ra_cluster][ra_name]]
                mentions_count = len(mentions)

                ra_cluster_name = ra_type+'_'+'_'.join(ra_cluster.split('_')[1:]) if ra_cluster != 'Unnamed' else ra_type+'_unnamed'

                # Append a row to the DataFrame
                row_data = [ra_cluster_name, ra_name, ra_type, mean_ra_score, owned, owned_percentage, mean_owned_score, reused, reused_percentage, mean_reused_score,
                            "\n".join([f"{license} ({percentage}%)" for license, percentage in all_licenses]),
                            "\n".join([f"{version} ({percentage}%)" for version, percentage in all_versions]),
                            "\n".join([f"{url} ({percentage}%)" for url, percentage in all_urls]),
                            "\n".join([f"{citation} ({percentage}%)" for citation, percentage in all_citations]),
                            mentions_count]

                row_series = pd.Series(row_data, index=columns)
                df = pd.concat([df, row_series.to_frame().T], ignore_index=True)

                # Add mentions to the mentions DataFrame
                for mention in mentions:
                    mentions_row_data = [mention[1], ra_cluster_name, ra_name, mention[2], mention[3], mention[4], mention[5], mention[6], mention[7], mention[8], mention[9], mention[10], mention[11], mention[12], mention[13], mention[14], mention[0]]
                    mentions_row_series = pd.Series(mentions_row_data, index=mentions_columns)
                    mentions_df = pd.concat([mentions_df, mentions_row_series.to_frame().T], ignore_index=True)

    # Save the reevaluated JSON to a file
    with open(file_path.replace('.json', '_reevaluated.json'), 'w') as f:
        json.dump({
            'publication_info': publication_info_df.to_dict()[''],
            'research_artifacts': df.to_dict('records'),
            'mentions': mentions_df.to_dict('records'),
            'citations': citations_df.to_dict('records'),
            'citances': citances_df.to_dict('records')
        }, f, indent=1)

    # Save the DataFrame to an Excel file
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        publication_info_df.to_excel(writer, header=False, sheet_name='Publication_Info')
        df.to_excel(writer, index=False, sheet_name='Research_Artifacts')
        mentions_df.to_excel(writer, index=False, sheet_name='Mentions')
        citations_df.to_excel(writer, index=False, sheet_name='Citations')
        citances_df.to_excel(writer, index=False, sheet_name='Citances')

    # Add hyperlinks to the "Mentions" sheet
    wb = load_workbook(output_file)
    ra_sheet = wb['Research_Artifacts']
    mentions_sheet = wb['Mentions']
    citations_sheet = wb['Citations']
    citances_sheet = wb['Citances']

    # For the 'Reseach_Artifacts' sheet, add a column in the beginning with the the 'Artifact ID' (e.g., 'A1', 'A2', etc.)
    ra_sheet.insert_cols(1)
    ra_sheet.cell(row=1, column=1, value='Artifact ID')  # Add column name
    ra_sheet.cell(row=1, column=1).font = Font(bold=True)
    for row_num in range(2, len(ra_sheet['B']) + 1):
        ra_sheet.cell(row=row_num, column=1, value=f'A{row_num - 1}')

    # Create a dictionary to store the row numbers of research artifacts in the "Research Artifacts" sheet
    ra_row_numbers = {}
    for row_num in range(2, len(df) + 2):
        ra_name = ra_sheet.cell(row=row_num, column=3).value
        ra_type = ra_sheet.cell(row=row_num, column=4).value
        ra_row_numbers[(ra_name, ra_type)] = row_num

    # Create a dictionary to store the FIRST row numbers of mentions in the "Mentions" sheet
    mentions_row_numbers = {}
    for row_num in range(2, len(mentions_df) + 2):
        ra_name = mentions_sheet.cell(row=row_num, column=3).value
        ra_type = mentions_sheet.cell(row=row_num, column=4).value
        ra_type = map_cand_type[ra_type]
        if (ra_name, ra_type) not in mentions_row_numbers:
            mentions_row_numbers[(ra_name, ra_type)] = row_num

    # Add hyperlinks to the "Mentions" sheet
    for row_num in range(2, len(mentions_df) + 2):
        ra_name = mentions_sheet.cell(row=row_num, column=3).value
        ra_type = mentions_sheet.cell(row=row_num, column=4).value
        ra_type = map_cand_type[ra_type]
        ra_row_num = ra_row_numbers.get((ra_name, ra_type), None)
        if ra_row_num:
            cell = mentions_sheet.cell(row=row_num, column=3)
            cell.hyperlink = f"#{ra_sheet.title}!A{ra_row_num}"
            cell.style = "Hyperlink"

    # Add hyperlinks to the "Research Artifacts" sheet
    for row_num in range(2, len(df) + 2):
        ra_name = ra_sheet.cell(row=row_num, column=3).value
        ra_type = ra_sheet.cell(row=row_num, column=4).value
        mentions_row_num = mentions_row_numbers.get((ra_name, ra_type), None)
        if mentions_row_num:
            cell = ra_sheet.cell(row=row_num, column=3)
            cell.hyperlink = f"#{mentions_sheet.title}!A{mentions_row_num}"
            cell.style = "Hyperlink"

    # Create a dictionary to store the row numbers of citations in the "Citations" sheet
    citations_row_numbers = {}
    for row_num in range(2, len(citations_df) + 2):
        citation_id = citations_sheet.cell(row=row_num, column=1).value
        citations_row_numbers[citation_id] = row_num

    # Create a dictionary to store the FIRST row numbers of citances in the "Citances" sheet
    citances_row_numbers = {}
    for row_num in range(2, len(citances_df) + 2):
        target_id = citances_sheet.cell(row=row_num, column=1).value
        if target_id and target_id[1:] not in citances_row_numbers:
            citances_row_numbers[target_id[1:]] = row_num

    # Add hyperlinks to the "Citances" sheet
    for row_num in range(2, len(citances) + 2):
        if citances_sheet.cell(row=row_num, column=1).value:
            target_id = citances_sheet.cell(row=row_num, column=1).value[1:]
            citation_row_num = citations_row_numbers.get(target_id, None)
            if citation_row_num:
                cell = citances_sheet.cell(row=row_num, column=1)
                cell.hyperlink = f"#{citations_sheet.title}!A{citation_row_num}"
                cell.style = "Hyperlink"

    # Add hyperlinks to the "Citations" sheet
    for row_num in range(2, len(citations_df) + 2):
        target_id = citations_sheet.cell(row=row_num, column=1).value
        citances_row_num = citances_row_numbers.get(target_id, None)
        if citances_row_num:
            cell = citations_sheet.cell(row=row_num, column=1)
            cell.hyperlink = f"#{citances_sheet.title}!A{citances_row_num}"
            cell.style = "Hyperlink"

    wb.save(output_file)
    wb.close()

    # Save another copy of the Excel file, excluding some of the details, so that it is easier to read and suitable to send to curators
    output_file_simple = file_path.replace('.json', '_reevaluated_simple.xlsx')

    # Load the Excel file with the hyperlinks
    wb = load_workbook(output_file)

    # The Excel has 4 sheets: 'Publication_Info', 'Research_Artifacts', 'Mentions', 'Citations', 'Citances'
    # Drop the following per sheet:
    # - 'Publication_Info': 'Claim', 'Claimer', 'Topic', 'FOS Pred 1', 'FOS Pred 2', 'FOS Pred 3' rows
    # - 'Research_Artifacts': 'Research Artifact Score', 'Owned Percentage', 'Owned Score', 'Reused Percentage', 'Reused Score' columns
    # - 'Mentions': 'Research Artifact Score', 'Owned Score', 'Reused Score', 'Indices' columns
    # - drop whole 'Citations' sheet
    # - drop whole 'Citances' sheet

    # Drop specified rows from 'Publication_Info' and columns from 'Research_Artifacts', 'Mentions' sheets
    sheets_to_modify = {
        'Publication_Info': ['Claim', 'Claimer', 'Topic', 'FOS Pred 1', 'FOS Pred 2', 'FOS Pred 3'],
        'Research_Artifacts': ['Research Artifact Score', 'Owned Percentage', 'Owned Score', 'Reused Percentage', 'Reused Score'],
        'Mentions': ['Research Artifact Score', 'Owned Score', 'Reused Score', 'Indices']
    }

    for sheet_name, items_to_drop in sheets_to_modify.items():
        sheet = wb[sheet_name]
        if sheet_name == 'Publication_Info':
            rows_to_drop = []
            for row in sheet.iter_rows():
                if row[0].value in items_to_drop:
                    rows_to_drop.append(row[0].row)
            for row_num in sorted(rows_to_drop, reverse=True):
                sheet.delete_rows(row_num)
        else:
            for column in items_to_drop:
                for cell in sheet[1]:
                    if cell.value == column:
                        sheet.delete_cols(cell.column)
                        break

    # Drop 'Citations' and 'Citances' sheets
    for sheet_name in ['Citations', 'Citances']:
        wb.remove(wb[sheet_name])
    
    # Remove the percentages
    ra_sheet = wb['Research_Artifacts']
    for row_num in range(2, len(ra_sheet['G']) + 1):
        prev_val = ra_sheet.cell(row=row_num, column=7).value
        if prev_val:
            new_val = "\n".join([f"{' '.join([x2 for x2 in x.split(' ')[:-1]])}" for x in prev_val.split('\n')])
            ra_sheet.cell(row=row_num, column=7, value=new_val)
        
        prev_val = ra_sheet.cell(row=row_num, column=8).value
        if prev_val:
            new_val = "\n".join([f"{' '.join([x2 for x2 in x.split(' ')[:-1]])}" for x in prev_val.split('\n')])
            ra_sheet.cell(row=row_num, column=8, value=new_val)
        
        prev_val = ra_sheet.cell(row=row_num, column=9).value
        if prev_val:
            new_val = "\n".join([f"{' '.join([x2 for x2 in x.split(' ')[:-1]])}" for x in prev_val.split('\n')])
            ra_sheet.cell(row=row_num, column=9, value=new_val)
        
        prev_val = ra_sheet.cell(row=row_num, column=10).value
        if prev_val:
            new_val = "\n".join([f"{' '.join([x2 for x2 in x.split(' ')[:-1]])}" for x in prev_val.split('\n')])
            ra_sheet.cell(row=row_num, column=10, value=new_val)

    # Convert the 'Type' in 'Mentions' tab through the map_cand_type in the Excel file
    mentions_sheet = wb['Mentions']
    for row_num in range(2, len(mentions_sheet['D']) + 1):
        mentions_sheet.cell(row=row_num, column=4, value=map_cand_type[mentions_sheet.cell(row=row_num, column=4).value])

    # Save the modified workbook
    wb.save(output_file_simple)
    wb.close()

""" 

*** ANALYZE IT MAIN  ****

"""

def extract_research_artifacts_pdf(input_path, xml_mode=False, filter_paragraphs=True, perform_deduplication=True, insert_fast_mode_gazetteers=False, dataset_gazetteers=None, verbose=True):
    # Check if the input is an XML file
    if xml_mode:
        # Parse the TEI XML + Extract citations and citances
        pdf_metadata = parse_tei_xml(input_path)
    else:
        # Parse the PDF + Extract citations and citances
        pdf_metadata = parse_pdf(input_path)

    if perform_deduplication:
        # Extract research artifacts
        candidates_metadata, grouped_clusters = extract_research_artifacts_pipeline(pdf_metadata, filter_paragraphs=filter_paragraphs, perform_deduplication=perform_deduplication, insert_fast_mode_gazetteers=insert_fast_mode_gazetteers, dataset_gazetteers=dataset_gazetteers, verbose=verbose)
    else:
        # Extract research artifacts
        candidates_metadata = extract_research_artifacts_pipeline(pdf_metadata, perform_deduplication=perform_deduplication, insert_fast_mode_gazetteers=insert_fast_mode_gazetteers, dataset_gazetteers=dataset_gazetteers, verbose=verbose)
        grouped_clusters = None

    # Output the results
    output = {
        'input_path': input_path,
        'pdf_metadata': pdf_metadata,
        'research_artifacts': {
            'candidates_metadata': candidates_metadata,
            'grouped_clusters': grouped_clusters
        }
    }

    return output


def extract_research_artifacts_pdf_fast_mode(input_path, xml_mode=False, dataset_gazetteers=None, verbose=True):
    # Check if the input is an XML file
    if xml_mode:
        # Parse the TEI XML + Extract citations and citances
        pdf_metadata = parse_tei_xml(input_path)
    else:
        # Parse the PDF + Extract citations and citances
        pdf_metadata = parse_pdf(input_path)

    research_artifacts = extract_research_artifacts_fast_mode(pdf_metadata, dataset_gazetteers=dataset_gazetteers, verbose=verbose)

    # Output the results
    output = {
        'input_path': input_path,
        'pdf_metadata': pdf_metadata,
        'research_artifacts': research_artifacts
    }

    return output

def extract_research_artifacts_parquet_file(input_path, filter_paragraphs=True, perform_deduplication=True, insert_fast_mode_gazetteers=False, dataset_gazetteers=None, verbose=True):
    # Load the parquet file
    parquet_data = pd.read_parquet(input_path)

    # Each parquet file contains multiple papers (each paper is a row)
    # so we need to iterate over each paper
    all_outputs = []
    if verbose:
        parquet_iter = tqdm(parquet_data.to_dict('records'))
    else:
        parquet_iter = parquet_data.to_dict('records')
    for row in parquet_iter:
        pdf_metadata = {
            'id': row['id'],
            'sections': eval(row['sections'])
        }
        if perform_deduplication:
            # Extract research artifacts
            candidates_metadata, grouped_clusters = extract_research_artifacts_pipeline(pdf_metadata, filter_paragraphs=filter_paragraphs, perform_deduplication=perform_deduplication, insert_fast_mode_gazetteers=insert_fast_mode_gazetteers, dataset_gazetteers=dataset_gazetteers, verbose=verbose)
        else:
            # Extract research artifacts
            candidates_metadata = extract_research_artifacts_pipeline(pdf_metadata, perform_deduplication=perform_deduplication, insert_fast_mode_gazetteers=insert_fast_mode_gazetteers, dataset_gazetteers=dataset_gazetteers, verbose=verbose)
            grouped_clusters = None

        # Output the results
        output = {
            'input_path': input_path,
            'pdf_metadata': pdf_metadata,
            'research_artifacts': {
                'candidates_metadata': candidates_metadata,
                'grouped_clusters': grouped_clusters
            }
        }

        # Add the output to the list of outputs
        all_outputs.append(output)

    return all_outputs


def extract_research_artifacts_parquet_file_fast_mode(input_path, dataset_gazetteers=None, verbose=True):
    # Load the parquet file
    parquet_data = pd.read_parquet(input_path)

    # Each parquet file contains multiple papers (each paper is a row)
    # so we need to iterate over each paper
    all_outputs = []
    if verbose:
        parquet_iter = tqdm(parquet_data.to_dict('records'))
    else:
        parquet_iter = parquet_data.to_dict('records')
    
    for row in parquet_iter:
        pdf_metadata = {
            'id': row['id'],
            'sections': eval(row['sections'])
        }
        research_artifacts = extract_research_artifacts_fast_mode(pdf_metadata, dataset_gazetteers=dataset_gazetteers, verbose=verbose)

        # Output the results
        output = {
            'input_path': input_path,
            'pdf_metadata': pdf_metadata,
            'research_artifacts': research_artifacts
        }

        # Add the output to the list of outputs
        all_outputs.append(output)

    return all_outputs


def extract_research_artifacts_parquet_file(input_path, filter_paragraphs=True, perform_deduplication=True, insert_fast_mode_gazetteers=False, dataset_gazetteers=None, verbose=True):
    # Load the parquet file
    parquet_data = pd.read_parquet(input_path)

    # Each parquet file contains multiple papers (each paper is a row)
    # so we need to iterate over each paper
    all_outputs = []
    if verbose:
        parquet_iter = tqdm(parquet_data.to_dict('records'))
    else:
        parquet_iter = parquet_data.to_dict('records')
    for row in parquet_iter:
        pdf_metadata = {
            'id': row['id'],
            'sections': eval(row['sections'])
        }
        if perform_deduplication:
            # Extract research artifacts
            candidates_metadata, grouped_clusters = extract_research_artifacts_pipeline(pdf_metadata, filter_paragraphs=filter_paragraphs, perform_deduplication=perform_deduplication, insert_fast_mode_gazetteers=insert_fast_mode_gazetteers, dataset_gazetteers=dataset_gazetteers, verbose=verbose)
        else:
            # Extract research artifacts
            candidates_metadata = extract_research_artifacts_pipeline(pdf_metadata, perform_deduplication=perform_deduplication, insert_fast_mode_gazetteers=insert_fast_mode_gazetteers, dataset_gazetteers=dataset_gazetteers, verbose=verbose)
            grouped_clusters = None

        # Output the results
        output = {
            'input_path': input_path,
            'pdf_metadata': pdf_metadata,
            'research_artifacts': {
                'candidates_metadata': candidates_metadata,
                'grouped_clusters': grouped_clusters
            }
        }

        # Add the output to the list of outputs
        all_outputs.append(output)

    return all_outputs


def extract_research_artifacts_parquet_file_fast_mode(input_path, dataset_gazetteers=None, verbose=True):
    # Load the parquet file
    parquet_data = pd.read_parquet(input_path)

    # Each parquet file contains multiple papers (each paper is a row)
    # so we need to iterate over each paper
    all_outputs = []
    if verbose:
        parquet_iter = tqdm(parquet_data.to_dict('records'))
    else:
        parquet_iter = parquet_data.to_dict('records')
    
    for row in parquet_iter:
        pdf_metadata = {
            'id': row['id'],
            'sections': eval(row['sections'])
        }
        research_artifacts = extract_research_artifacts_fast_mode(pdf_metadata, dataset_gazetteers=dataset_gazetteers, verbose=verbose)

        # Output the results
        output = {
            'input_path': input_path,
            'pdf_metadata': pdf_metadata,
            'research_artifacts': research_artifacts
        }

        # Add the output to the list of outputs
        all_outputs.append(output)

    return all_outputs


def extract_research_artifacts_text_file(input_path, split_sentences=False, perform_deduplication=True, insert_fast_mode_gazetteers=False, dataset_gazetteers=None, verbose=True):

    with open(input_path, 'r') as f:
        text_list = f.readlines()
    
    # Split sentences if needed using the NLTK tokenizer
    if split_sentences:
        text_list = [sent_tokenize(x) for x in text_list]
    else:
        text_list = [text_list]

    if perform_deduplication:
        # Extract research artifacts
        candidates_metadata, grouped_clusters = extract_research_artifacts_list(text_list, perform_deduplication=perform_deduplication, insert_fast_mode_gazetteers=insert_fast_mode_gazetteers, dataset_gazetteers=dataset_gazetteers, verbose=verbose)
    else:
        # Extract research artifacts
        candidates_metadata = extract_research_artifacts_list(text_list, perform_deduplication=perform_deduplication, insert_fast_mode_gazetteers=insert_fast_mode_gazetteers, dataset_gazetteers=dataset_gazetteers, verbose=verbose)
        grouped_clusters = None

    # Output the results
    output = {
        'input_path': input_path,
        'text_list': text_list,
        'research_artifacts': {
            'candidates_metadata': candidates_metadata,
            'grouped_clusters': grouped_clusters
        }
    }

    return output


def extract_research_artifacts_text_file_fast_mode(input_path, split_sentences=False, dataset_gazetteers=None, verbose=True):

    with open(input_path, 'r') as f:
        text_list = f.readlines()

    # Split sentences if needed using the NLTK tokenizer
    if split_sentences:
        text_list = [sent_tokenize(x) for x in text_list]
    else:
        text_list = [text_list]

    research_artifacts = extract_research_artifacts_list_fast_mode(text_list, dataset_gazetteers=dataset_gazetteers, verbose=verbose)

    # Output the results
    output = {
        'input_path': input_path,
        'text_list': text_list,
        'research_artifacts': research_artifacts
    }

    return output


def extract_research_artifacts_text_list(text_list, split_sentences=False, perform_deduplication=True, insert_fast_mode_gazetteers=False, dataset_gazetteers=None, verbose=True):

    # Split sentences if needed using the NLTK tokenizer
    if split_sentences:
        text_list = [[i for s in [sent_tokenize(sent) for sent in par] for i in s] for par in text_list]

    if perform_deduplication:
        # Extract research artifacts
        candidates_metadata, grouped_clusters = extract_research_artifacts_list(text_list, perform_deduplication=perform_deduplication, insert_fast_mode_gazetteers=insert_fast_mode_gazetteers, dataset_gazetteers=dataset_gazetteers, verbose=verbose)
    else:
        # Extract research artifacts
        candidates_metadata = extract_research_artifacts_list(text_list, perform_deduplication=perform_deduplication, insert_fast_mode_gazetteers=insert_fast_mode_gazetteers, dataset_gazetteers=dataset_gazetteers, verbose=verbose)
        grouped_clusters = None

    # Output the results
    output = {
        'text_list': text_list,
        'research_artifacts': {
            'candidates_metadata': candidates_metadata,
            'grouped_clusters': grouped_clusters
        }
    }

    return output


def extract_research_artifacts_text_list_fast_mode(text_list, split_sentences=False, dataset_gazetteers=None, verbose=True):

    # Split sentences if needed using the NLTK tokenizer
    if split_sentences:
        text_list = [[i for s in [sent_tokenize(sent) for sent in par] for i in s] for par in text_list]

    research_artifacts = extract_research_artifacts_list_fast_mode(text_list, dataset_gazetteers=dataset_gazetteers, verbose=verbose)

    # Output the results
    output = {
        'text_list': text_list,
        'research_artifacts': research_artifacts
    }

    return output


def main():
    parser = argparse.ArgumentParser(description='Bulk inference of citances polarity and intent from PDF files.')
    parser.add_argument('--input_dir', type=str, help='Directory with PDF files to process.', required=True)
    parser.add_argument('--output_dir', type=str, help='Output directory to save the results.', required=True)
    parser.add_argument('--filter_paragraphs', action='store_true', help='Filter paragraphs to only include the ones with citations.')
    parser.add_argument('--perform_deduplication', action='store_true', help='Perform deduplication of the mentions.')
    parser.add_argument('--insert_fast_mode_gazetteers', action='store_true', help='Insert the fast mode gazetteers.')
    parser.add_argument('--dataset_gazetteers', type=str, choices=['hybrid', 'synthetic', None], help='Type of dataset gazetteers to use (synthetic, hybrid).')
    parser.add_argument('--fast_mode', action='store_true', help='Use the fast mode to extract research artifacts.')
    parser.add_argument('--text_mode', action='store_true', help='Extract research artifacts from text files instead of pdfs. Supersedes fast_mode and perform_deduplication options.')
    parser.add_argument('--text_mode_split_sentences', action='store_true', help='Split the text into sentences.')
    parser.add_argument('--xml_mode', action='store_true', help='Run the pipeline for publication PDFs that have been already been processed by GROBID. Requires input dir to contain TEI XML files instead of PDFs.')
    parser.add_argument('--parquet_mode', action='store_true', help='Run the pipeline for parquet files instead of PDFs that contain the columns: "id", "sections" .')
    parser.add_argument('--reevaluate', action='store_true', help='Reevaluate the results using new thresholds.')
    parser.add_argument('--reevaluate_only', action='store_true', help='Only reevaluate the results using new thresholds, skipping the normal pipeline.')
    parser.add_argument('--filter_input', type=str, help='Wildcard pattern to filter input files to analyze.')
    parser.add_argument('--thresholds', type=str, help='The thresholds to use in the reevaluation. Valid keys are valid_score, ownership_score, and usage_score.')
    parser.add_argument('--verbose', action='store_true', help='Verbose mode.')


    args = parser.parse_args()

    # Create the output directory if it does not exist
    if not os.path.exists(args.output_dir):
        os.makedirs(args.output_dir)

    if not args.reevaluate_only:
        if args.text_mode:
            # Get the list of text files
            print(f'Processing text files in {args.input_dir}')
            text_files = [f for f in os.listdir(args.input_dir) if f.endswith('.txt')]
            if args.filter_input:
                text_files = fnmatch.filter(text_files, args.filter_input)

            for text_file in text_files:
                print(f'Processing {text_file}')
                text_path = os.path.join(args.input_dir, text_file)
                output_path = os.path.join(args.output_dir, text_file.replace('.txt', '.json'))

                if args.fast_mode:
                    res = extract_research_artifacts_text_file_fast_mode(text_path, args.text_mode_split_sentences, args.dataset_gazetteers, args.verbose)
                else:
                    res = extract_research_artifacts_text_file(text_path, args.text_mode_split_sentences, args.perform_deduplication, args.insert_fast_mode_gazetteers, args.dataset_gazetteers, args.verbose)

                with open(output_path, 'w') as f:
                    json.dump(res, f, indent=1)
            
            print('Finished processing all text files. Output saved in:', args.output_dir)
        elif args.parquet_mode:
            # Get the list of parquet files
            print(f'Processing parquet files in {args.input_dir}')
            parquet_files = [f for f in os.listdir(args.input_dir) if f.endswith('.parquet')]
            if args.filter_input:
                parquet_files = fnmatch.filter(parquet_files, args.filter_input)

            for parquet_file in parquet_files:
                print(f'Processing {parquet_file}')
                parquet_path = os.path.join(args.input_dir, parquet_file)
                output_path = os.path.join(args.output_dir, parquet_file.replace('.parquet', '.json'))

                if args.fast_mode:
                    res = extract_research_artifacts_parquet_file_fast_mode(parquet_path, args.dataset_gazetteers, args.verbose)
                else:
                    res = extract_research_artifacts_parquet_file(parquet_path, args.filter_paragraphs, args.perform_deduplication, args.insert_fast_mode_gazetteers, args.dataset_gazetteers, args.verbose)

                with open(output_path, 'w') as f:
                    json.dump(res, f, indent=1)
            
            print('Finished processing all parquet files. Output saved in:', args.output_dir)
        else:
            # Get the list of PDF/XML files
            if args.xml_mode:
                print(f'Processing TEI XML files in {args.input_dir}')
                pdf_files = [f for f in os.listdir(args.input_dir) if f.endswith('.xml')]
            else:
                print(f'Processing PDF files in {args.input_dir}')
                pdf_files = [f for f in os.listdir(args.input_dir) if f.endswith('.pdf')]

            if args.filter_input:
                pdf_files = fnmatch.filter(pdf_files, args.filter_input)

            for pdf_file in pdf_files:
                print(f'Processing {pdf_file}')
                pdf_path = os.path.join(args.input_dir, pdf_file)
                if args.xml_mode:
                    output_path = os.path.join(args.output_dir, pdf_file.replace('.xml', '.json'))
                else:
                    output_path = os.path.join(args.output_dir, pdf_file.replace('.pdf', '.json'))

                if args.fast_mode:
                    res = extract_research_artifacts_pdf_fast_mode(pdf_path, args.xml_mode, args.dataset_gazetteers, args.verbose)
                else:
                    res = extract_research_artifacts_pdf(pdf_path, args.xml_mode, args.filter_paragraphs, args.perform_deduplication, args.insert_fast_mode_gazetteers, args.dataset_gazetteers, args.verbose)

                with open(output_path, 'w') as f:
                    json.dump(res, f, indent=1)
            
            if args.xml_mode:
                print('Finished processing all TEI XML files. Output saved in:', args.output_dir)
            else:
                print('Finished processing all PDF files. Output saved in:', args.output_dir)

    if args.reevaluate or args.reevaluate_only:
        if args.text_mode or args.parquet_mode:
            print('Reevaluation works only in PDF/XML mode. Skipping reevaluation.')
        else:
            # Convert the thresholds string back to a dictionary
            if args.thresholds:
                arg_thresholds = json.loads(args.thresholds)
            else:
                arg_thresholds = {}
            
            new_thresholds = {
                'artifact_answer': arg_thresholds['valid_score'] if 'valid_score' in arg_thresholds else 0.8,
                'ownership_answer': arg_thresholds['ownership_score'] if 'ownership_score' in arg_thresholds else 0.8,
                'reuse_answer': arg_thresholds['usage_score'] if 'usage_score' in arg_thresholds else 0.8
            }

            print('Reevaluating the results using new thresholds...')
            res_files = [f for f in os.listdir(args.output_dir) if f.endswith('.json')]
            if args.filter_input:
                res_files = fnmatch.filter(res_files, args.filter_input)

            for res_file in res_files:
                print(f'Reevaluating {res_file}')
                try:
                    reevaluate(os.path.join(args.output_dir, res_file), new_thresholds)
                except Exception as e:
                    print(f'Error reevaluating {res_file}, is it a valid PDF/XML output file? Reevaluation works only in PDF/XML mode.')
            
            print('Finished reevaluating all results. Output saved in:', args.output_dir)


if __name__ == '__main__':
    main()
